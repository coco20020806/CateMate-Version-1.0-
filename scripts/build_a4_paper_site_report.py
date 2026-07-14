"""Build Part1/Part2 HTML report and data workbook from 植印A4纸.xlsx."""

from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SOURCE = Path(r"C:\Users\fei.kong\Desktop\daily\植印A4纸.xlsx")
OUTPUT_DIR = Path(r"C:\Users\fei.kong\Desktop\daily\植印A4纸_report")
CR_LABELS = ["CR5", "CR10", "CR20"]
CR_ORDER_ROWS = (44, 45, 46)
CR_GMV_SKU_ROWS = (72, 73, 74)
CR_GMV_SHOP_ROWS = (82, 83, 84)
PRICE_TIER_ROWS = range(105, 122)
TOP_ORDERS_ROWS = range(22, 42)
TOP_GMV_ROWS = range(50, 70)
Y26_SHEET_NAMES = ("Y26", "Sheet10")
L3_SHEET_NAMES = ("L3",)
GMV_SOURCE_SHEET_NAMES = ("gmv source",)
GMV_SOURCE_ROW_START = 1
GMV_SOURCE_ROW_END = 9
GMV_SOURCE_COL_START = 1
GMV_SOURCE_COL_END = 7
SITE_ORDER = ["BR", "ID", "TH", "VN", "PH", "MY", "TW", "SG"]
TREND_METRICS = [
    ("orders", "Orders", "Part5_Trend_Orders"),
    ("orders_pct", "Orders Site %", "Part5_Trend_Orders_SitePct"),
    ("gmv_usd", "GMV (USD)", "Part5_Trend_GMV_USD"),
    ("gmv_pct", "GMV Site %", "Part5_Trend_GMV_SitePct"),
    ("aov", "客单价 (USD)", "Part5_Trend_AOV"),
]
PART5_DEFAULT_TREND_METRIC = "orders"
PART1_ADG_GROWTH_METRIC = "gmv_usd"
PART1_ADG_AOV_GROWTH_METRIC = "aov"
TREND_COL_MAP = {
    "orders": 2,
    "orders_pct": 3,
    "gmv_usd": 4,
    "gmv_pct": 5,
    "aov": 6,
}
SOURCE_SHEET_SKIP = {"Sheet10", "Y26", "L3", "gmv source"}

BRAND_PATTERNS = [
    ("IK Yellow / Indah Kiat", r"ik\s*yellow|indah\s*kiat|ik\s*黄色|ik黄"),
    ("Paper One", r"paper\s*one|paperone"),
    ("Double A", r"double\s*a|ดั๊บเบิ้ล\s*เอ"),
    ("Deli", r"\bdeli\b"),
    ("FCI", r"\bfci\b"),
    ("APRIL", r"\bapril\b"),
    ("IK Eco", r"ik\s*eco"),
    ("UP COPY", r"up\s*copy"),
]

PRODUCT_PATTERNS = [
    ("复印纸 / 打印纸", r"copier|copy\s*paper|复印|打印纸|bond\s*paper|sulfite|kertas|กระดาษถ่าย"),
    ("相纸 / 照片纸", r"photo\s*paper|fotogr|glossy|matte|相纸|照片纸|相紙"),
    ("打印服务", r"printing\s*-|print\s*service|打印\s*-|photocopy|按页|/页"),
    ("卡纸 / 艺术纸", r"art\s*card|card\s*paper|卡纸|象牙"),
    ("贴纸", r"sticker|贴纸|stiker"),
    ("彩色纸", r"colour\s*paper|color\s*paper|彩色纸|warna"),
]

BUNDLE_PATTERNS = [
    ("整箱 / 5令装", r"5\s*ream|5\s*令|box|carton|kotak|กล่อง"),
    ("大包装 / 多令", r"10\s*x|2500|10\s*ream|整箱"),
    ("单令", r"1\s*ream|单令|1\s*令"),
]


def load_workbook():
    return openpyxl.load_workbook(SOURCE, data_only=True)


def resolve_sheet_name(wb, candidates: tuple[str, ...]) -> str | None:
    for name in candidates:
        if name in wb.sheetnames:
            return name
    return None


def site_sheet_names(wb) -> list[str]:
    return [name for name in SITE_ORDER if name in wb.sheetnames]


def format_month_label(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if value is None:
        return ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return text


def parse_monthly_site_blocks(ws) -> list[dict]:
    blocks = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value != "Date" or ws.cell(row, 2).value in (None, ""):
            continue
        month = format_month_label(ws.cell(row, 2).value)
        sites: dict[str, dict[str, object]] = {}
        for data_row in range(row + 2, row + 12):
            site = ws.cell(data_row, 1).value
            if not site or site == "TTL":
                continue
            sites[str(site)] = {
                "orders": ws.cell(data_row, 2).value,
                "orders_pct": ws.cell(data_row, 3).value,
                "gmv_usd": ws.cell(data_row, 4).value,
                "gmv_pct": ws.cell(data_row, 5).value,
                "aov": ws.cell(data_row, 6).value,
            }
        blocks.append({"month": month, "sites": sites})
    blocks.sort(key=lambda item: item["month"])
    return blocks


def parse_l3_wide_sheet(ws) -> list[dict]:
    """Parse L3 wide layout: months in row 1 (B+), sites in column A (row 2+)."""
    month_cols: list[tuple[int, str]] = []
    for col in range(2, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header is None:
            continue
        month_cols.append((col, format_month_label(header)))

    blocks = [{"month": month, "sites": {}} for _, month in month_cols]
    month_index = {month: idx for idx, (_, month) in enumerate(month_cols)}

    for row in range(2, ws.max_row + 1):
        site = ws.cell(row, 1).value
        if not site:
            continue
        site_name = str(site)
        for col, month in month_cols:
            value = ws.cell(row, col).value
            if value is None:
                continue
            idx = month_index[month]
            blocks[idx]["sites"][site_name] = {"gmv_usd": value}

    blocks.sort(key=lambda item: item["month"])
    return blocks


def parse_l3_sheet(ws) -> list[dict]:
    first_cell = ws.cell(1, 1).value
    if first_cell == "Date":
        return parse_monthly_site_blocks(ws)
    return parse_l3_wide_sheet(ws)


def blocks_to_gmv_pivot(blocks: list[dict]) -> dict:
    months = [block["month"] for block in blocks]
    sites = ordered_sites(blocks)
    return {
        "months": months,
        "sites": sites,
        "series": {
            site: [
                block["sites"].get(site, {}).get("gmv_usd")
                for block in blocks
            ]
            for site in sites
        },
    }


def ordered_sites(blocks: list[dict]) -> list[str]:
    found = {site for block in blocks for site in block["sites"]}
    ordered = [site for site in SITE_ORDER if site in found]
    ordered.extend(sorted(found - set(ordered)))
    return ordered


def build_trend_tables(blocks: list[dict]) -> dict:
    months = [block["month"] for block in blocks]
    sites = ordered_sites(blocks)
    tables = {}
    for metric_key, _, _ in TREND_METRICS:
        tables[metric_key] = {
            "months": months,
            "sites": sites,
            "series": {
                site: [
                    block["sites"].get(site, {}).get(metric_key)
                    for block in blocks
                ]
                for site in sites
            },
        }
    return tables


def build_a4_l3_ratio(y26_blocks: list[dict], l3_blocks: list[dict]) -> dict | None:
    if not y26_blocks or not l3_blocks:
        return None

    l3_lookup = {
        (block["month"], site): block["sites"].get(site, {}).get("gmv_usd")
        for block in l3_blocks
        for site in block["sites"]
    }
    months = [block["month"] for block in y26_blocks]
    sites = ordered_sites(y26_blocks)
    series = {}
    for site in sites:
        values = []
        for block in y26_blocks:
            a4_gmv = block["sites"].get(site, {}).get("gmv_usd")
            l3_gmv = l3_lookup.get((block["month"], site))
            if a4_gmv is None or l3_gmv in (None, 0):
                values.append(None)
            else:
                values.append(round(float(a4_gmv) / float(l3_gmv), 6))
        series[site] = values

    return {"months": months, "sites": sites, "series": series}


def prepare_trend_chart_data(trend_tables: dict) -> dict:
    pct_metrics = {"orders_pct", "gmv_pct"}
    chart_data = {}
    for metric_key, table in trend_tables.items():
        series = {}
        for site, values in table["series"].items():
            series[site] = [
                pct_to_chart(v)
                if metric_key in pct_metrics and v is not None
                else (float(v) if v is not None else None)
                for v in values
            ]
        chart_data[metric_key] = {
            "months": table["months"],
            "sites": table["sites"],
            "series": series,
        }
    return chart_data


def calc_growth_pct(current, base) -> float | None:
    if current is None or base in (None, 0):
        return None
    return round((float(current) / float(base) - 1) * 100, 2)


def build_trend_growth_tables(trend_tables: dict) -> dict:
    growth_tables = {}
    for metric_key, table in trend_tables.items():
        months = table["months"]
        sites = table["sites"]
        tabs = []
        if len(months) >= 2:
            tabs.append(
                {
                    "key": "h1",
                    "label": "半年增速",
                    "subtitle": f"{months[0]} → {months[-1]}",
                    "values": [
                        calc_growth_pct(
                            table["series"][site][-1],
                            table["series"][site][0],
                        )
                        for site in sites
                    ],
                }
            )
        for idx in range(1, len(months)):
            tabs.append(
                {
                    "key": months[idx],
                    "label": f"{months[idx]} MoM",
                    "subtitle": f"较 {months[idx - 1]}",
                    "values": [
                        calc_growth_pct(
                            table["series"][site][idx],
                            table["series"][site][idx - 1],
                        )
                        for site in sites
                    ],
                }
            )
        growth_tables[metric_key] = {"sites": sites, "tabs": tabs}
    return growth_tables


def prepare_trend_growth_chart_data(growth_tables: dict) -> dict:
    return growth_tables


def prepare_ratio_chart_data(a4_l3_ratio: dict | None) -> dict | None:
    if not a4_l3_ratio:
        return None
    series = {
        site: [
            pct_to_chart(v) if v is not None else None for v in values
        ]
        for site, values in a4_l3_ratio["series"].items()
    }
    return {
        "months": a4_l3_ratio["months"],
        "sites": a4_l3_ratio["sites"],
        "series": series,
    }


def write_trend_workbook_sheets(
    out_wb, trend_tables, a4_l3_ratio, l3_gmv_pivot=None, growth_tables=None
):
    for metric_key, _, sheet_name in TREND_METRICS:
        table = trend_tables[metric_key]
        ws = out_wb.create_sheet(sheet_name)
        ws.append(["month"] + table["sites"])
        for idx, month in enumerate(table["months"]):
            row = [month]
            for site in table["sites"]:
                row.append(table["series"][site][idx])
            ws.append(row)

    if growth_tables:
        growth_sheet_names = {
            "orders": "Part5_Growth_Orders",
            "orders_pct": "Part5_Growth_Orders_SitePct",
            "gmv_usd": "Part5_Growth_GMV_USD",
            "gmv_pct": "Part5_Growth_GMV_SitePct",
            "aov": "Part5_Growth_AOV",
        }
        for metric_key, growth in growth_tables.items():
            ws = out_wb.create_sheet(growth_sheet_names[metric_key])
            ws.append(["tab", "note"] + growth["sites"])
            for tab in growth["tabs"]:
                ws.append([tab["label"], tab["subtitle"]] + tab["values"])

    if l3_gmv_pivot:
        ws = out_wb.create_sheet("Part5_L3_GMV")
        ws.append(["month"] + l3_gmv_pivot["sites"])
        for idx, month in enumerate(l3_gmv_pivot["months"]):
            row = [month]
            for site in l3_gmv_pivot["sites"]:
                row.append(l3_gmv_pivot["series"][site][idx])
            ws.append(row)

    ws = out_wb.create_sheet("Part5_A4_GMV_in_L3_Ratio")
    ws.append(["month"] + (a4_l3_ratio["sites"] if a4_l3_ratio else []))
    if a4_l3_ratio:
        for idx, month in enumerate(a4_l3_ratio["months"]):
            row = [month]
            for site in a4_l3_ratio["sites"]:
                row.append(a4_l3_ratio["series"][site][idx])
            ws.append(row)


def read_part1(ws):
    sites, orders, orders_pct, gmv_usd, gmv_pct, aov = [], [], [], [], [], []
    for row in range(5, 13):
        sites.append(ws.cell(row, 2).value)
        orders.append(ws.cell(row, 3).value)
        orders_pct.append(ws.cell(row, 4).value)
        gmv_usd.append(ws.cell(row, 5).value)
        gmv_pct.append(ws.cell(row, 6).value)
        aov.append(ws.cell(row, 7).value)
    return {
        "sites": sites,
        "orders": orders,
        "orders_pct": orders_pct,
        "gmv_usd": gmv_usd,
        "gmv_pct": gmv_pct,
        "aov": aov,
    }


def format_aov_usd(value) -> str:
    if value is None:
        return "-"
    return f"{round(float(value), 1):.1f}"


def format_gmv_usd_k(value) -> str:
    if value is None:
        return "-"
    return f"{round(float(value) / 1000):,}K"


def format_growth_band(value) -> str:
    if value is None:
        return "-"
    low = math.floor(float(value) / 10) * 10
    high = low + 10
    return f"{int(low)}%-{int(high)}%"


def growth_band_midpoint(value) -> float | None:
    if value is None:
        return None
    low = math.floor(float(value) / 10) * 10
    return (low + (low + 10)) / 2


def h1_growth_by_site(growth_tables: dict | None, metric_key: str) -> dict[str, float | None]:
    if not growth_tables or metric_key not in growth_tables:
        return {}
    growth = growth_tables[metric_key]
    h1_tab = next((tab for tab in growth["tabs"] if tab["key"] == "h1"), None)
    if not h1_tab:
        return {}
    return dict(zip(growth["sites"], h1_tab["values"]))


def build_part1_adg_table_html(
    part1: dict, growth_tables: dict | None = None
) -> str:
    h1_gmv_growth = h1_growth_by_site(growth_tables, PART1_ADG_GROWTH_METRIC)
    h1_aov_growth = h1_growth_by_site(growth_tables, PART1_ADG_AOV_GROWTH_METRIC)
    gmv_label = next(
        label for key, label, _ in TREND_METRICS if key == PART1_ADG_GROWTH_METRIC
    )
    aov_label = next(
        label for key, label, _ in TREND_METRICS if key == PART1_ADG_AOV_GROWTH_METRIC
    )
    rows = []
    for i, site in enumerate(part1["sites"]):
        rows.append(
            f"<tr>"
            f"<td>{site}</td>"
            f"<td>{format_gmv_usd_k(part1['gmv_usd'][i])}</td>"
            f"<td>{format_aov_usd(part1['aov'][i])}</td>"
            f"<td>{format_growth_band(h1_gmv_growth.get(site))}</td>"
            f"<td>{format_growth_band(h1_aov_growth.get(site))}</td>"
            f"</tr>"
        )
    h1_subtitle = ""
    if growth_tables and PART1_ADG_GROWTH_METRIC in growth_tables:
        h1_tab = next(
            (
                tab
                for tab in growth_tables[PART1_ADG_GROWTH_METRIC]["tabs"]
                if tab["key"] == "h1"
            ),
            None,
        )
        if h1_tab:
            h1_subtitle = h1_tab["subtitle"]
    return f"""
    <div class="card wide">
      <h3>各 Site ADG 量级与客单价</h3>
      <table class="sku-table part1-adg-table">
        <thead>
          <tr>
            <th>Site</th>
            <th>ADG 量级 (K USD)</th>
            <th>客单价 (USD)</th>
            <th>半年增速区间 ({gmv_label})</th>
            <th>半年增速区间 ({aov_label})</th>
          </tr>
        </thead>
        <tbody>
          {"".join(rows)}
        </tbody>
      </table>
      <p class="chart-footnote">
        客单价 = GMV ÷ Orders，保留 1 位小数。
        半年增速区间按 10 个百分点分档展示（如 0%-10%、10%-20%）：{gmv_label} 与 {aov_label} 均为半年增速（{h1_subtitle or "2026-01 → 2026-06"}），与 Part5 增速区间图一致。
        <br>数据时间范围：2026年6月（单月）；增速基期见 Part5（Y26 月度数据）。
      </p>
    </div>
    """


def _gmv_source_to_pct(value) -> float:
    if value in (None, ""):
        return 0.0
    fv = float(value)
    if abs(fv) <= 1.5:
        return round(fv * 100, 2)
    return round(fv, 2)


def read_gmv_source(wb) -> dict | None:
    sheet_name = resolve_sheet_name(wb, GMV_SOURCE_SHEET_NAMES)
    if not sheet_name:
        return None

    ws = wb[sheet_name]
    source_names: list[str] = []
    for col in range(GMV_SOURCE_COL_START + 1, GMV_SOURCE_COL_END + 1):
        val = ws.cell(GMV_SOURCE_ROW_START, col).value
        if val is None or str(val).strip() == "":
            break
        source_names.append(str(val).strip())
    if not source_names:
        return None

    data_rows: dict[str, list[float]] = {}
    for row in range(GMV_SOURCE_ROW_START + 1, GMV_SOURCE_ROW_END + 1):
        site_val = ws.cell(row, GMV_SOURCE_COL_START).value
        if site_val is None or str(site_val).strip() == "":
            continue
        site = str(site_val).strip().upper()
        vals = [
            _gmv_source_to_pct(ws.cell(row, GMV_SOURCE_COL_START + 1 + idx).value)
            for idx in range(len(source_names))
        ]
        data_rows[site] = vals

    if not data_rows:
        return None

    sites = [s for s in SITE_ORDER if s in data_rows] + [
        s for s in data_rows if s not in SITE_ORDER
    ]
    series = {
        name: [data_rows[s][i] for s in sites]
        for i, name in enumerate(source_names)
    }
    stack_totals = [
        round(sum(data_rows[s][i] for i in range(len(source_names))), 2)
        for s in sites
    ]
    return {
        "sites": sites,
        "sources": source_names,
        "series": series,
        "stack_totals": stack_totals,
        "period_note": (
            "数据时间范围：2026年7月 MTD。"
            "Top20 Shop 排名口径：仅按各 Site 店铺 A4 item sum 6月 GMV 排序取前 20 名；"
            "图中展示上述 Top20 店铺各 ADG 来源占比。"
            "百分比分母为 total adgmv，由于数据统计因素总数不一定为 100%。"
            "数值轴已脱敏；Site 名称与来源图例正常显示。"
        ),
    }


def read_cr_block(ws, rows: tuple[int, ...], cr_col: int):
    values = []
    for row in rows:
        cr = ws.cell(row, cr_col).value
        values.append(cr if cr is not None else 0)
    return values


def read_price_tiers(ws):
    tiers = []
    for row in PRICE_TIER_ROWS:
        tiers.append(
            {
                "tier": ws.cell(row, 1).value,
                "gmv": ws.cell(row, 2).value,
                "gmv_pct": ws.cell(row, 3).value,
                "orders": ws.cell(row, 5).value,
                "orders_pct": ws.cell(row, 6).value,
            }
        )
    return tiers


def pct_to_chart(v):
    if v is None:
        return 0
    return round(float(v) * 100, 4)


def read_top_skus(ws):
    def read_block(rows):
        items = []
        for row in rows:
            rank = ws.cell(row, 1).value
            if rank is None or not isinstance(rank, (int, float)):
                continue
            items.append(
                {
                    "rank": int(rank),
                    "price_range": ws.cell(row, 5).value or "",
                    "item_name": ws.cell(row, 6).value or "",
                    "item_link": ws.cell(row, 7).value or "",
                    "translation": ws.cell(row, 8).value or "",
                }
            )
        return items

    return {
        "by_orders": read_block(TOP_ORDERS_ROWS),
        "by_gmv": read_block(TOP_GMV_ROWS),
    }


def _match_patterns(text: str, patterns: list[tuple[str, str]]) -> list[str]:
    found = []
    for label, pattern in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            found.append(label)
    return found


def _top_counter(values: list[str], limit: int = 3) -> list[str]:
    ranked = [item for item, _ in Counter(values).most_common(limit)]
    return ranked


def extract_site_profile(site: str, top_skus: dict) -> dict:
    all_items = top_skus["by_orders"] + top_skus["by_gmv"]
    if not all_items:
        return {
            "site": site,
            "brands": [],
            "products": [],
            "gsm": [],
            "tiers": [],
            "bundles": [],
            "has_a4b5": False,
            "has_print_service": False,
            "has_photo_paper": False,
        }

    combined_text = " ".join(
        f"{item['item_name']} {item['translation']}" for item in all_items
    )
    lower_text = combined_text.lower()

    brands = _top_counter(_match_patterns(lower_text, BRAND_PATTERNS), limit=4)
    products = _top_counter(_match_patterns(lower_text, PRODUCT_PATTERNS), limit=3)
    bundles = _match_patterns(lower_text, BUNDLE_PATTERNS)

    gsm_values = re.findall(r"(\d{2,3})\s*g(?:sm)?", lower_text, re.IGNORECASE)
    gsm_labels = [f"{g}gsm" for g in _top_counter(gsm_values)[:3]]

    tier_values = [
        item["price_range"] for item in all_items if item.get("price_range")
    ]
    tier_ranked = _top_counter(tier_values, limit=3)

    has_photo_paper = bool(
        re.search(r"photo\s*paper|fotogr|glossy|相纸|照片纸|相紙", lower_text)
    )
    has_print_service = bool(
        re.search(r"printing\s*-|print\s*service|打印\s*-|photocopy|按页|/页", lower_text)
    )
    has_a4b5 = bool(re.search(r"a4/b5|b5", lower_text))

    return {
        "site": site,
        "brands": brands,
        "products": products,
        "gsm": gsm_labels,
        "tiers": tier_ranked,
        "bundles": bundles,
        "has_a4b5": has_a4b5,
        "has_print_service": has_print_service,
        "has_photo_paper": has_photo_paper,
    }


def _format_site_summary(profile: dict) -> str:
    site = profile["site"]
    parts = [f"**{site} 市场 Top SKU 特征总结（基于 Rank by Orders / GMV Top 20 商品名）：**"]

    if profile["brands"]:
        parts.append(
            f"- **品牌：** 高频出现 {'、'.join(profile['brands'])}，头部 SKU 品牌集中度较明显。"
        )
    if profile["products"]:
        parts.append(f"- **品类：** 以 {'、'.join(profile['products'])} 为主。")
    if profile["gsm"]:
        parts.append(f"- **规格：** 常见克重为 {'、'.join(profile['gsm'])}。")
    if profile["bundles"]:
        parts.append(
            f"- **包装：** 可见 {'、'.join(profile['bundles'][:2])} 等售卖形式。"
        )
    if profile["tiers"]:
        parts.append(
            f"- **价格带：** Top SKU 多集中在 {'、'.join(profile['tiers'])}。"
        )
    if profile["has_a4b5"]:
        parts.append("- **尺寸：** 除 A4 外，部分商品以 A4/B5 组合或关联规格售卖。")
    if profile["has_print_service"]:
        parts.append(
            "- **服务类 SKU：** 存在按页计费的打印/复印类商品，反映本地按需打印需求。"
        )

    if len(parts) == 1:
        parts.append("- Top SKU 以 A4 相关办公用纸为主，规格与包装形式较为多元。")

    return "\n".join(parts)


def generate_site_summary(site: str, top_skus: dict) -> str:
    return _format_site_summary(extract_site_profile(site, top_skus))


def generate_cross_site_summary(sites: list[str], top_skus: dict) -> str:
    profiles = [extract_site_profile(site, top_skus[site]) for site in sites]

    parts = [
        "**各 Site 横向概括（2026年6月 Top SKU）：**",
        "以下基于各市场 Rank by Orders / GMV Top 20 商品名称，对各 Site 头部 SKU 结构进行横向比较。",
    ]

    brand_map: dict[str, list[str]] = defaultdict(list)
    for profile in profiles:
        for brand in profile["brands"][:2]:
            if profile["site"] not in brand_map[brand]:
                brand_map[brand].append(profile["site"])

    if brand_map:
        brand_lines = [
            f"{brand}（{'、'.join(site_list)}）"
            for brand, site_list in sorted(
                brand_map.items(), key=lambda item: (-len(item[1]), item[0])
            )[:5]
        ]
        parts.append(f"- **品牌格局：** {'；'.join(brand_lines)}。")

    copier_sites = [
        p["site"] for p in profiles if "复印纸 / 打印纸" in p["products"]
    ]
    photo_sites = [p["site"] for p in profiles if p["has_photo_paper"]]
    print_sites = [p["site"] for p in profiles if p["has_print_service"]]

    if copier_sites:
        parts.append(
            f"- **共性：** {'、'.join(copier_sites)} 等市场的 Top SKU 均以复印纸 / 打印纸为主，A4 办公用纸需求占主导。"
        )
    if photo_sites:
        parts.append(
            f"- **相纸结构：** {'、'.join(photo_sites)} 的 Top SKU 中相纸 / 照片纸占比较高，与其他 Site 形成差异。"
        )
    if print_sites:
        parts.append(
            f"- **文印服务：** {'、'.join(print_sites)} 可见按页打印 / 复印类 SKU，体现本地化按需文印需求。"
        )

    parts.append("- **分市场概览：**")
    for profile in profiles:
        highlights = []
        if profile["brands"]:
            highlights.append(f"头部品牌偏 {'、'.join(profile['brands'][:2])}")
        if profile["products"]:
            highlights.append(f"品类以 {'、'.join(profile['products'][:2])} 为主")
        if profile["gsm"]:
            highlights.append(f"常见规格 {'、'.join(profile['gsm'][:2])}")
        if profile["tiers"]:
            highlights.append(f"价格带集中在 {'、'.join(profile['tiers'][:2])}")
        if not highlights:
            highlights.append("以 A4 相关办公用纸为主")
        parts.append(f"  - **{profile['site']}：** {'；'.join(highlights)}。")

    return "\n".join(parts)


def build_data_workbook(
    wb,
    part1,
    cr_orders,
    cr_gmv_sku,
    cr_gmv_shop,
    price_tiers,
    top_skus,
    summaries,
    cross_site_summary,
    trend_tables,
    a4_l3_ratio,
    l3_gmv_pivot=None,
    growth_tables=None,
):
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    ws1 = out_wb.create_sheet("Part1_各市场分布")
    ws1.append(["site", "orders", "orders占比", "gmv_usd", "gmv占比", "客单价(usd)", "gmv半年增速区间", "客单价半年增速区间"])
    h1_gmv_growth = h1_growth_by_site(growth_tables, PART1_ADG_GROWTH_METRIC)
    h1_aov_growth = h1_growth_by_site(growth_tables, PART1_ADG_AOV_GROWTH_METRIC)
    for i, site in enumerate(part1["sites"]):
        ws1.append(
            [
                site,
                part1["orders"][i],
                part1["orders_pct"][i],
                part1["gmv_usd"][i],
                part1["gmv_pct"][i],
                round(float(part1["aov"][i]), 1) if part1["aov"][i] is not None else None,
                format_growth_band(h1_gmv_growth.get(site)),
                format_growth_band(h1_aov_growth.get(site)),
            ]
        )

    def write_cr_sheet(name, data):
        ws = out_wb.create_sheet(name)
        ws.append(["site"] + CR_LABELS)
        for site, values in data.items():
            ws.append([site] + values)

    write_cr_sheet("Part2_CR_Orders", cr_orders)
    write_cr_sheet("Part2_CR_GMV_SKU", cr_gmv_sku)
    write_cr_sheet("Part2_CR_GMV_Shop", cr_gmv_shop)

    tier_names = [t["tier"] for t in price_tiers[next(iter(price_tiers))]]
    sites = list(price_tiers.keys())

    for sheet_name, key in [
        ("Part2_PriceTier_GMV", "gmv"),
        ("Part2_PriceTier_GMV占比", "gmv_pct"),
        ("Part2_PriceTier_Orders", "orders"),
        ("Part2_PriceTier_Orders占比", "orders_pct"),
    ]:
        ws = out_wb.create_sheet(sheet_name)
        ws.append(["Price tier"] + sites)
        for tier in tier_names:
            row = [tier]
            for site in sites:
                match = next(
                    (t for t in price_tiers[site] if t["tier"] == tier),
                    None,
                )
                row.append(match[key] if match else None)
            ws.append(row)

    sku_headers = [
        "site",
        "rank",
        "price_range",
        "item_name",
        "translation",
        "item_link",
    ]
    for sheet_name, key in [
        ("Part4_TopSKU_ByOrders", "by_orders"),
        ("Part4_TopSKU_ByGMV", "by_gmv"),
    ]:
        ws = out_wb.create_sheet(sheet_name)
        ws.append(sku_headers)
        for site, blocks in top_skus.items():
            for item in blocks[key]:
                ws.append(
                    [
                        site,
                        item["rank"],
                        item["price_range"],
                        item["item_name"],
                        item["translation"],
                        item["item_link"],
                    ]
                )

    ws_summary = out_wb.create_sheet("Part4_Site_Summary")
    ws_summary.append(["site", "summary"])
    ws_summary.append(["ALL_SITES", cross_site_summary])
    for site, summary in summaries.items():
        ws_summary.append([site, summary])

    write_trend_workbook_sheets(
        out_wb, trend_tables, a4_l3_ratio, l3_gmv_pivot, growth_tables
    )

    for ws in out_wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

    xlsx_path = OUTPUT_DIR / "植印A4纸_数据底表.xlsx"
    out_wb.save(xlsx_path)
    return xlsx_path


def format_tier_label(tier: str) -> str:
    """Format tier label for matrix row header."""
    if not tier or not str(tier).startswith("Tier"):
        return str(tier)
    parts = str(tier).split("_", 1)
    if len(parts) != 2:
        return str(tier)
    num = parts[0].replace("Tier", "").zfill(2)
    band = parts[1].replace("-", ",")
    if band.endswith("+"):
        band = band[:-1] + ",Inf)"
    else:
        band = f"[{band})"
    return f"{num}_{band}"


def _metric_totals_by_month(table: dict) -> list[float]:
    totals = []
    for idx in range(len(table["months"])):
        values = [
            table["series"][site][idx]
            for site in table["sites"]
            if table["series"][site][idx] is not None
        ]
        totals.append(float(sum(values)))
    return totals


def _fmt_int(value: float) -> str:
    return f"{int(round(value)):,}"


def _fmt_pct(value: float, digits: int = 1) -> str:
    return f"{value:.{digits}f}%"


def _desensitize_growth(value: float) -> str:
    if abs(value) < 2:
        return "基本持平"
    rounded = round(value / 5) * 5
    if rounded == 0:
        return "基本持平"
    return f"约 {abs(rounded)}%{'增长' if value > 0 else '下降'}"


def _desensitize_volume(value: float) -> str:
    if value >= 1_000_000:
        millions = round(value / 500_000) * 0.5
        if millions.is_integer():
            return f"约 {int(millions)} 百万"
        return f"约 {millions} 百万"
    rounded = round(value / 50_000) * 5
    return f"约 {rounded} 万"


def _desensitize_share(pct: float) -> str:
    if pct < 1:
        return "约 1% 以内"
    return f"约 {round(pct / 5) * 5}%"


def _desensitize_ratio_range(low: float, high: float) -> str:
    lo = round(low / 5) * 5
    hi = round(high / 5) * 5
    if lo == hi:
        return f"约 {lo}%"
    return f"约 {lo}%–{hi}%"


def build_executive_summary(
    part1: dict,
    trend_tables: dict,
    a4_l3_ratio: dict | None,
    l3_gmv_pivot: dict | None,
) -> str:
    order_totals = _metric_totals_by_month(trend_tables["orders"])
    gmv_totals = _metric_totals_by_month(trend_tables["gmv_usd"])

    orders_growth = (order_totals[-1] / order_totals[0] - 1) * 100
    gmv_growth = (gmv_totals[-1] / gmv_totals[0] - 1) * 100

    aov_start = gmv_totals[0] / order_totals[0]
    aov_end = gmv_totals[-1] / order_totals[-1]
    aov_change = (aov_end / aov_start - 1) * 100

    top_order_sites = sorted(
        zip(part1["sites"], part1["orders_pct"]),
        key=lambda item: pct_to_chart(item[1]),
        reverse=True,
    )[:3]
    top_order_text = "、".join(
        f"{site}（{_desensitize_share(pct_to_chart(pct))}）"
        for site, pct in top_order_sites
    )

    ratio_band = "约 70%–80%"
    if a4_l3_ratio and a4_l3_ratio["months"]:
        ratio_values = []
        for idx in range(len(a4_l3_ratio["months"])):
            vals = [
                pct_to_chart(a4_l3_ratio["series"][site][idx])
                for site in a4_l3_ratio["sites"]
                if a4_l3_ratio["series"][site][idx] is not None
            ]
            if vals:
                ratio_values.append(sum(vals) / len(vals))
        if ratio_values:
            ratio_band = _desensitize_ratio_range(min(ratio_values), max(ratio_values))

    l3_trend_text = "文具纸类母类目整体仍有一定扩张"
    if l3_gmv_pivot and len(l3_gmv_pivot["months"]) >= 2:
        l3_totals = _metric_totals_by_month(l3_gmv_pivot)
        l3_2025 = [
            total
            for month, total in zip(l3_gmv_pivot["months"], l3_totals)
            if str(month).startswith("2025")
        ]
        l3_2026 = [
            total
            for month, total in zip(l3_gmv_pivot["months"], l3_totals)
            if str(month).startswith("2026")
        ]
        if l3_2025 and l3_2026:
            l3_growth = (
                (sum(l3_2026) / len(l3_2026)) / (sum(l3_2025) / len(l3_2025)) - 1
            ) * 100
            l3_trend_text = (
                f"文具纸类母类目规模{_desensitize_growth(l3_growth)}，"
                "但 A4 细分增速相对温和"
            )

    orders_growth_text = _desensitize_growth(orders_growth)
    gmv_growth_text = _desensitize_growth(gmv_growth)
    aov_change_text = _desensitize_growth(aov_change)
    volume_range = (
        f"{_desensitize_volume(order_totals[0])} 至 {_desensitize_volume(order_totals[-1])}"
    )

    return f"""
    <section class="executive-summary" id="executive-summary">
      <h2>卖家导读 · A4 纸市场概览</h2>
      <div class="verdict-card">
        <div class="verdict-label">给卖家的一句话</div>
        <div class="verdict-title">成熟标品市场，有需求、有竞争，靠差异化经营取胜</div>
        <p class="verdict-desc">
          如果您正在评估是否进入或加大 A4 纸布局：这个类目<strong>不是短期爆发型风口</strong>，但<strong>需求稳定、市场体量可观</strong>。
          上半年平台订单规模从 {volume_range} 区间波动上行，整体呈温和扩张。
          更适合有供应链、品牌或规格卡位能力的卖家，以<strong>长期经营</strong>而非投机铺货来切入。
        </p>
      </div>

      <div class="summary-grid">
        <div class="summary-kpi">
          <div class="kpi-label">上半年订单趋势</div>
          <div class="kpi-value">{orders_growth_text}</div>
          <div class="kpi-sub">全站点合计，月度间有起伏</div>
        </div>
        <div class="summary-kpi">
          <div class="kpi-label">上半年 GMV 趋势</div>
          <div class="kpi-value">{gmv_growth_text}</div>
          <div class="kpi-sub">规模维持增长，节奏不平滑</div>
        </div>
        <div class="summary-kpi">
          <div class="kpi-label">客单价走势</div>
          <div class="kpi-value">{aov_change_text}</div>
          <div class="kpi-sub">提价空间有限，不宜只靠涨价</div>
        </div>
        <div class="summary-kpi">
          <div class="kpi-label">A4 在纸类类目地位</div>
          <div class="kpi-value">{ratio_band}</div>
          <div class="kpi-sub">A4 仍是纸类核心需求，但份额高位趋稳</div>
        </div>
      </div>

      <div class="summary-sections">
        <div class="summary-block">
          <h3>一、市场大环境：卖家需要知道什么？</h3>
          <ul>
            <li><strong>需求还在，但别指望躺赢：</strong>上半年订单与 GMV 均呈温和上行（分别{orders_growth_text}、{gmv_growth_text}），说明买家需求仍在，但月度波动明显，备货与投放需按「稳态+波动」来规划。</li>
            <li><strong>客单价难拉高：</strong>整体客单价{aov_change_text}，买家对价格敏感，单靠提价很难拉动 GMV，更应从规格组合、箱装/令装、复购装入手。</li>
            <li><strong>纸类大盘仍在长大：</strong>{l3_trend_text}。A4 占纸类 GMV {ratio_band}，已是主力品类——意味着<strong>市场够大</strong>，但也意味着<strong>竞争集中在这个主战场</strong>。</li>
            <li><strong>商品高度同质化：</strong>头部链接以复印纸/打印纸为主，常见品牌包括 IK Yellow、Paper One、Double A 等，70/80gsm、单令/整箱是主流规格，中低价格带（Tier 3–5）最拥挤。</li>
          </ul>
        </div>
        <div class="summary-block">
          <h3>二、各市场怎么选？（站点视角）</h3>
          <ul>
            <li><strong>优先关注大盘 Site：</strong>当前订单量主要集中在 {top_order_text} 等市场，适合有成本与履约能力、能跑规模的卖家优先布局。</li>
            <li><strong>高客单小众市场：</strong>TW、SG 客单价明显高于均值，适合推高端规格、组合装或品牌溢价，但绝对体量较小，更适合作为利润补充而非唯一阵地。</li>
            <li><strong>差异化机会：</strong>部分市场（如 MY）可见打印服务、组合规格类 SKU 上榜，说明本地化场景（复印、装订、按页打印）仍能切出细分需求。</li>
            <li><strong>竞争格局：</strong>多数市场 CR20 处于中等水平——头部链接能拿到可观份额，但长尾仍有空间；小市场（如 SG）集中度更高，新进入者需更精准选品。</li>
          </ul>
        </div>
        <div class="summary-block">
          <h3>三、给卖家的实操建议</h3>
          <ul>
            <li><strong>定位：</strong>按成熟标品运营——重点放在<strong>品牌/白牌认知、克重规格、令装箱装、价格带卡位</strong>，而非期待品类级爆发红利。</li>
            <li><strong>选品：</strong>优先覆盖主流 70/80gsm 复印纸，并用箱装、多令组合提升客单；关注各 Site 头部价格带差异，避免全市场同一定价。</li>
            <li><strong>竞争：</strong>研究各市场 Top SKU 的标题与规格写法（见 Part4），对标头部但不盲目低价内卷；MY 等市场可探索「纸+服务」组合。</li>
            <li><strong>节奏：</strong>关注月度趋势与 A4/纸类占比变化——若占比持续走平而纸类大盘仍涨，说明增量在向其他纸类细分迁移，需及时调整 SKU 结构。</li>
          </ul>
        </div>
      </div>
      <p class="chart-footnote">本导读面向卖家视角，数据已做脱敏处理；详细图表与明细见下方各 Part。</p>
    </section>
    """


def build_html(
    part1,
    cr_orders,
    cr_gmv_sku,
    cr_gmv_shop,
    price_tiers,
    top_skus,
    summaries,
    cross_site_summary,
    trend_tables,
    a4_l3_ratio,
    html_name: str,
    l3_gmv_pivot=None,
    gmv_source=None,
    growth_tables=None,
):
    executive_summary_html = build_executive_summary(
        part1, trend_tables, a4_l3_ratio, l3_gmv_pivot
    )
    part1_adg_table_html = build_part1_adg_table_html(part1, growth_tables)
    sites = part1["sites"]
    cr_sites = list(cr_orders.keys())
    tier_names = [t["tier"] for t in price_tiers[next(iter(price_tiers))]]
    tier_labels = [format_tier_label(t) for t in tier_names]

    chart_data = {
        "part1": {
            "sites": sites,
            "orders_pct": [pct_to_chart(v) for v in part1["orders_pct"]],
            "gmv_pct": [pct_to_chart(v) for v in part1["gmv_pct"]],
            "aov": [float(v) if v is not None else 0 for v in part1["aov"]],
        },
        "cr_orders": {
            "sites": cr_sites,
            "series": {
                label: [pct_to_chart(cr_orders[s][i]) for s in cr_sites]
                for i, label in enumerate(CR_LABELS)
            },
        },
        "cr_gmv_sku": {
            "sites": cr_sites,
            "series": {
                label: [pct_to_chart(cr_gmv_sku[s][i]) for s in cr_sites]
                for i, label in enumerate(CR_LABELS)
            },
        },
        "cr_gmv_shop": {
            "sites": cr_sites,
            "series": {
                label: [pct_to_chart(cr_gmv_shop[s][i]) for s in cr_sites]
                for i, label in enumerate(CR_LABELS)
            },
        },
        "price_tier": {
            "sites": cr_sites,
            "tiers": tier_labels,
            "gmv_pct": {
                site: [pct_to_chart(t["gmv_pct"]) for t in price_tiers[site]]
                for site in cr_sites
            },
            "orders_pct": {
                site: [pct_to_chart(t["orders_pct"]) for t in price_tiers[site]]
                for site in cr_sites
            },
        },
        "top_skus": top_skus,
        "summaries": summaries,
        "cross_site_summary": cross_site_summary,
        "trend": prepare_trend_chart_data(trend_tables),
        "trend_growth": prepare_trend_growth_chart_data(growth_tables or {}),
        "a4_l3_ratio": prepare_ratio_chart_data(a4_l3_ratio),
        "has_l3_ratio": a4_l3_ratio is not None,
        "gmv_source": gmv_source,
        "has_gmv_source": gmv_source is not None,
    }

    data_json = json.dumps(chart_data, ensure_ascii=False)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>植印A4纸 - 各市场分析报告</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
  <style>
    :root {{
      --bg: #f7f5f2;
      --card: #ffffff;
      --text: #4a4a48;
      --muted: #8a8580;
      --border: #e8e4df;
      --accent: #b8956b;
      --orange: #d4a574;
      --orange-light: #e8c9a8;
      --orange-dark: #b07d4f;
      --bar-light: #e8c9a8;
      --bar-dark: #b07d4f;
      --matrix-banner: #a67b6b;
    }}
    html {{
      scroll-behavior: smooth;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.5;
    }}
    .container {{ max-width: 1280px; margin: 0 auto; padding: 24px; }}
    header {{
      background: linear-gradient(135deg, #e8ecf0, #ede9e4);
      color: #111111;
      padding: 32px 24px;
      border-radius: 12px;
      margin-bottom: 24px;
    }}
    header h1 {{ margin: 0 0 8px; font-size: 28px; color: #111111; }}
    header p {{ margin: 0; color: #111111; opacity: 0.85; }}
    h1, h2, h3 {{
      color: #111111;
    }}
    h2 {{
      font-size: 22px;
      margin: 32px 0 16px;
      padding-bottom: 8px;
      border-bottom: 2px solid var(--accent);
      scroll-margin-top: 20px;
    }}
    h3 {{ font-size: 16px; margin: 0 0 12px; font-weight: 600; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(340px, 1fr));
      gap: 20px;
    }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 20px;
      box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }}
    .card.wide {{ grid-column: 1 / -1; }}
    .chart-wrap {{ position: relative; height: 360px; }}
    .note, .section-note, .chart-footnote {{
      font-size: 13px;
      color: var(--muted);
      margin-top: 8px;
      line-height: 1.6;
    }}
    .section-note {{
      margin: -8px 0 16px;
      padding: 10px 14px;
      background: #f5efe8;
      border-left: 3px solid var(--accent);
      border-radius: 4px;
      color: #7a6a58;
    }}
    footer {{ text-align: center; color: var(--muted); font-size: 13px; margin: 32px 0 8px; }}

    .tier-matrix {{
      border: 1px solid #d1d5db;
      background: #fff;
      overflow-x: auto;
    }}
    .matrix-banner {{
      background: #f0ebe6;
      color: #111111;
      text-align: center;
      font-size: 18px;
      font-weight: 700;
      padding: 10px 12px;
      letter-spacing: 0.5px;
      border-bottom: 2px solid var(--accent);
    }}
    .tier-grid {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      font-size: 12px;
    }}
    .tier-grid th,
    .tier-grid td {{
      border: 1px dotted #b8c4d6;
      padding: 0;
      vertical-align: middle;
    }}
    .tier-grid thead th {{
      background: #f8fafc;
      font-weight: 700;
      text-align: center;
      padding: 8px 4px;
      font-size: 13px;
      color: #111111;
    }}
    .tier-grid .corner {{
      text-align: left;
      padding-left: 10px;
      width: 110px;
    }}
    .tier-label {{
      padding: 6px 8px;
      white-space: nowrap;
      font-size: 11px;
      color: #374151;
      width: 110px;
    }}
    .bar-cell {{
      height: 28px;
      padding: 4px 6px;
    }}
    .bar-track {{
      width: 100%;
      height: 100%;
      display: flex;
      align-items: center;
    }}
    .bar {{
      height: 14px;
      min-width: 0;
      border-radius: 1px;
    }}
    .bar-normal {{ background: var(--bar-light); }}
    .bar-max {{ background: var(--bar-dark); }}
    .matrix-footnote {{
      text-align: right;
      font-size: 11px;
      color: var(--muted);
      padding: 6px 10px;
      border-top: 1px dotted #b8c4d6;
    }}

    .site-tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 16px;
    }}
    .site-tab {{
      border: 1px solid var(--border);
      background: #fff;
      color: var(--text);
      padding: 8px 16px;
      border-radius: 20px;
      cursor: pointer;
      font-size: 13px;
      font-weight: 600;
      transition: all 0.15s ease;
    }}
    .site-tab:hover {{ border-color: var(--orange); color: var(--orange-dark); }}
    .site-tab.active {{
      background: var(--orange);
      border-color: var(--orange);
      color: #fff;
    }}
    .metric-tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 16px;
    }}
    .metric-tab {{
      border: 1px solid var(--border);
      background: #fff;
      color: var(--text);
      padding: 8px 14px;
      border-radius: 20px;
      cursor: pointer;
      font-size: 13px;
      font-weight: 600;
      transition: all 0.15s ease;
    }}
    .metric-tab:hover {{ border-color: var(--orange); color: var(--orange-dark); }}
    .metric-tab.active {{
      background: var(--orange);
      border-color: var(--orange);
      color: #fff;
    }}
    .summary-box {{
      background: #f5efe8;
      border-left: 3px solid var(--accent);
      border-radius: 6px;
      padding: 14px 16px;
      margin-bottom: 18px;
      font-size: 14px;
      line-height: 1.7;
      white-space: pre-wrap;
    }}
    .summary-box.overview {{
      background: #f0f4f3;
      border-left-color: #9aabbd;
      margin-bottom: 14px;
    }}
    .sku-block {{ margin-bottom: 24px; }}
    .sku-table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }}
    .sku-table th,
    .sku-table td {{
      border: 1px solid var(--border);
      padding: 10px 12px;
      vertical-align: top;
      text-align: left;
    }}
    .sku-table th {{
      background: #f8f6f3;
      font-weight: 600;
      color: #5c574f;
    }}
    .sku-table tr:nth-child(even) td {{ background: #fcfbfa; }}
    .sku-table .rank-col {{ width: 52px; text-align: center; }}
    .sku-table .tier-col {{ width: 110px; white-space: nowrap; }}
    .part1-adg-table th:nth-child(2),
    .part1-adg-table td:nth-child(2),
    .part1-adg-table th:nth-child(3),
    .part1-adg-table td:nth-child(3),
    .part1-adg-table th:nth-child(4),
    .part1-adg-table td:nth-child(4),
    .part1-adg-table th:nth-child(5),
    .part1-adg-table td:nth-child(5) {{
      text-align: right;
      white-space: nowrap;
    }}
    .sku-table a {{ color: #9e7b57; word-break: break-all; }}
    .name-cn {{ font-weight: 600; }}
    .name-foreign {{ color: var(--muted); margin-top: 4px; font-size: 12px; }}

    .executive-summary {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 24px;
      margin-bottom: 28px;
      box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }}
    .executive-summary > h2 {{
      margin-top: 0;
      border-bottom-color: #9aabbd;
    }}
    .verdict-card {{
      background: linear-gradient(135deg, #f5efe8, #f0f4f3);
      border-left: 4px solid var(--orange-dark);
      border-radius: 8px;
      padding: 18px 20px;
      margin-bottom: 20px;
    }}
    .verdict-label {{
      font-size: 12px;
      letter-spacing: 1px;
      text-transform: uppercase;
      color: var(--muted);
      margin-bottom: 6px;
    }}
    .verdict-title {{
      font-size: 24px;
      font-weight: 700;
      color: #111111;
      margin-bottom: 10px;
    }}
    .verdict-desc {{ margin: 0; line-height: 1.8; }}
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 14px;
      margin-bottom: 22px;
    }}
    .summary-kpi {{
      background: #faf9f7;
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 14px;
    }}
    .kpi-label {{ font-size: 12px; color: var(--muted); margin-bottom: 6px; }}
    .kpi-value {{ font-size: 22px; font-weight: 700; color: var(--orange-dark); }}
    .kpi-sub {{ font-size: 12px; color: var(--muted); margin-top: 4px; }}
    .summary-sections {{
      display: grid;
      gap: 16px;
    }}
    .summary-block {{
      background: #fcfbfa;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 16px 18px;
    }}
    .summary-block h3 {{
      margin: 0 0 10px;
      font-size: 15px;
      color: #111111;
    }}
    .summary-block ul {{
      margin: 0;
      padding-left: 20px;
      line-height: 1.8;
      font-size: 14px;
    }}
    .report-nav {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 24px;
      margin: 0 0 28px;
      box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }}
    .report-nav h2 {{
      margin-top: 0;
      font-size: 18px;
      border-bottom: none;
      padding-bottom: 0;
      margin-bottom: 14px;
    }}
    .nav-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 10px;
    }}
    .nav-link {{
      display: block;
      padding: 12px 14px;
      border: 1px solid var(--border);
      border-radius: 8px;
      text-decoration: none;
      color: var(--text);
      background: #faf9f7;
      font-size: 14px;
      transition: all 0.15s ease;
    }}
    .nav-link:hover {{
      border-color: var(--orange);
      background: #fff8f0;
      color: var(--orange-dark);
    }}
    .nav-link .nav-part {{
      display: block;
      font-size: 11px;
      color: var(--muted);
      margin-bottom: 4px;
      letter-spacing: 0.5px;
    }}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1>植印A4纸 · 各市场分析报告</h1>
      <p>数据源：植印A4纸.xlsx · 数据时间范围：2026年6月（单月） · 生成时间：{generated_at}</p>
    </header>

    {executive_summary_html}

    <nav class="report-nav" id="report-nav">
      <h2>报告目录 · 快速跳转</h2>
      <div class="nav-grid">
        <a class="nav-link" href="#executive-summary"><span class="nav-part">摘要</span>卖家导读 · A4 纸市场概览</a>
        <a class="nav-link" href="#part1"><span class="nav-part">Part1</span>各市场 orders / GMV 分布情况</a>
        <a class="nav-link" href="#part2"><span class="nav-part">Part2</span>市场集中度（CR5 / CR10 / CR20）</a>
        <a class="nav-link" href="#part3"><span class="nav-part">Part3</span>Price Tier 分布</a>
        <a class="nav-link" href="#part4"><span class="nav-part">Part4</span>Top SKU 明细</a>
        <a class="nav-link" href="#part5-trend"><span class="nav-part">Part5</span>A4 纸月度趋势（By Site）</a>
        <a class="nav-link" href="#part6-ratio"><span class="nav-part">Part6</span>A4 纸 GMV 占 L3 品类比例趋势</a>
        <a class="nav-link" href="#part7-gmv-source"><span class="nav-part">Part7</span>Top20 Shop ADG 来源占比</a>
      </div>
    </nav>

    <h2 id="part1">Part1：各市场 orders / GMV 分布情况</h2>
    <p class="section-note">筛选范围：Item name 含 A4 的相关 item。饼图展示各 Site 在上述 A4 item 范围内的 Orders / GMV 占比。</p>
    <div class="grid">
      <div class="card">
        <h3>各 Site Orders 占比</h3>
        <div class="chart-wrap"><canvas id="ordersPie"></canvas></div>
        <p class="chart-footnote">各 Site A4 item Orders 占全部 Site A4 item Orders 总和的比例。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="card">
        <h3>各 Site GMV 占比</h3>
        <div class="chart-wrap"><canvas id="gmvPie"></canvas></div>
        <p class="chart-footnote">各 Site A4 item GMV 占全部 Site A4 item GMV 总和的比例。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="card wide">
        <h3>各 Site 客单价</h3>
        <div class="chart-wrap"><canvas id="aovBar"></canvas></div>
        <p class="chart-footnote">各 Site A4 item 客单价（GMV / Orders）。数值轴已脱敏；Site 名称正常显示。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
    </div>
    {part1_adg_table_html}

    <h2 id="part2">Part2：市场集中度</h2>
    <div class="grid">
      <div class="card wide">
        <h3>Top Items Rank by Orders — CR5 / CR10 / CR20</h3>
        <div class="chart-wrap"><canvas id="crOrders"></canvas></div>
        <p class="chart-footnote">CR 含义：该 Site 当月按 Orders 排名前 N 的 item，其 Orders 加总占该 Site 总 Orders 的比例（N = 5 / 10 / 20）。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="card wide">
        <h3>Top Items Rank by GMV — CR5 / CR10 / CR20</h3>
        <div class="chart-wrap"><canvas id="crGmvSku"></canvas></div>
        <p class="chart-footnote">CR 含义：该 Site 当月按 GMV 排名前 N 的 item，其 GMV 加总占该 Site 总 GMV 的比例（N = 5 / 10 / 20）。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="card wide">
        <h3>Top Shops Rank by GMV (Only Counts A4 SKU) — CR5 / CR10 / CR20</h3>
        <div class="chart-wrap"><canvas id="crGmvShop"></canvas></div>
        <p class="chart-footnote">CR 含义：该 Site 当月按 GMV 排名前 N 的 shop，其 A4 SKU GMV 加总占该 Site A4 SKU 总 GMV 的比例（N = 5 / 10 / 20）。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
    </div>

    <h2 id="part3">Part3：Price Tier 分布</h2>
    <div class="grid">
      <div class="card wide">
        <div id="tierGmvMatrix"></div>
        <p class="chart-footnote">各 Site 内各价格段 GMV 占比；柱长为该 Site 内相对占比，最高价格段以深橙标注。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="card wide">
        <div id="tierOrdersMatrix"></div>
        <p class="chart-footnote">各 Site 内各价格段 Orders 占比；柱长为该 Site 内相对占比，最高价格段以深橙标注。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
    </div>

    <h2 id="part4">Part4：Top SKU 明细</h2>
    <p class="section-note">数据时间范围：2026年6月（单月）。点击不同 Site 切换查看该市场 Rank by Orders / GMV 的 Top 20 SKU，含中文名、外文名、价格带与商品链接；总结文字基于 Top SKU 名称归纳各 Site 的品牌、规格与市场特点。</p>
    <div class="card wide">
      <div class="summary-box overview" id="crossSiteSummary"></div>
      <div class="site-tabs" id="siteTabs"></div>
      <div class="summary-box" id="siteSummary"></div>
      <div class="sku-block">
        <h3>Rank by Orders — Top 20 SKU</h3>
        <div id="skuOrdersTable"></div>
        <p class="chart-footnote">数据来源：各 Site sheet A19:H41（sort by orders）。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
      <div class="sku-block">
        <h3>Rank by GMV — Top 20 SKU</h3>
        <div id="skuGmvTable"></div>
        <p class="chart-footnote">数据来源：各 Site sheet A48:H69（sort by GMV）。<br>数据时间范围：2026年6月（单月）。</p>
      </div>
    </div>

    <h2 id="part5-trend">Part5：A4 纸月度趋势（By Site）</h2>
    <p class="section-note">数据来源：Y26 sheet（各月 A4 纸 by Site 汇总），对应数据底表 <code>Part5_Trend_*</code> 与 <code>Part5_Growth_*</code>。上方折线图可切换 Orders / GMV 等五项指标；下方两张增速柱状图结构一致：默认<strong>半年增速</strong>，可切换各月 <strong>MoM</strong>；柱顶标注为 10 个百分点增速区间（如 0%-10%、10%-20%）。第一张随折线图指标 Tab 变化，第二张固定 GMV (USD)。</p>
    <div class="card wide">
      <h3>月度趋势（折线图）</h3>
      <div class="metric-tabs" id="trendMetricTabs"></div>
      <div class="chart-wrap"><canvas id="trendLineChart"></canvas></div>
      <p class="chart-footnote" id="trendFootnote">数值轴已脱敏；时间轴与 Site 图例正常显示。</p>
    </div>
    <div class="card wide">
      <h3>各 Site Orders 增速区间（柱状图）</h3>
      <div class="metric-tabs" id="growthPeriodTabs"></div>
      <div class="chart-wrap"><canvas id="trendGrowthBar"></canvas></div>
      <p class="chart-footnote" id="trendGrowthFootnote">默认展示 Orders 半年增速区间；切换上方折线图指标 Tab 后同步切换指标，周期 Tab 可选各月 MoM。</p>
    </div>
    <div class="card wide">
      <h3>各 Site GMV 增速区间（柱状图）</h3>
      <div class="metric-tabs" id="gmvGrowthPeriodTabs"></div>
      <div class="chart-wrap"><canvas id="trendGmvBandBar"></canvas></div>
      <p class="chart-footnote" id="trendGmvBandFootnote">GMV (USD) 增速区间；默认半年增速，可切换各月 MoM。</p>
    </div>

    <h2 id="part6-ratio">Part6：A4 纸 GMV 占 L3 品类比例趋势</h2>
    <p class="section-note">A4 纸 GMV 取自 Y26（D 列），L3 品类 GMV 取自 L3 sheet（Printing &amp; Photocopy Paper 所属 L3 月度 by Site 汇总）。比例为：当月该 Site A4 GMV ÷ 当月该 Site L3 GMV。</p>
    <div class="card wide">
      <div class="chart-wrap"><canvas id="ratioLineChart"></canvas></div>
      <p class="chart-footnote" id="ratioFootnote">数值轴已脱敏；时间轴与 Site 图例正常显示。</p>
      <p class="chart-footnote" id="ratioMissingNote" style="display:none;color:#b45309;">当前数据源中未找到 L3 sheet，无法生成 A4/L3 占比趋势图；请将 L3 数据补充至 Excel 后重新生成报告。</p>
    </div>

    <h2 id="part7-gmv-source">Part7：Top20 Shop ADG 来源占比（堆积柱状图）</h2>
    <p class="section-note">数据来源：植印A4纸.xlsx · sheet <code>gmv source</code>（A1:G9）。横轴为 Site，纵轴为各 ADG 来源占比（%）；每一根柱为该 Site <strong>Top20 Shop</strong> 的 ADG 来源结构。<br><strong>数据时间范围：2026年7月 MTD。</strong> Top20 Shop 排名口径：仅按各 Site 店铺 <strong>A4 item sum 6月 GMV</strong> 排序取前 20 名。百分比分母为 <strong>total adgmv</strong>，由于数据统计因素总数不一定为 100%。</p>
    <div class="card wide">
      <h3>各 Site Top20 Shop ADG 来源结构</h3>
      <div class="chart-wrap" style="height:420px"><canvas id="gmvSourceStacked"></canvas></div>
      <p class="chart-footnote" id="gmvSourceFootnote">数值轴已脱敏；Site 名称与来源图例正常显示。</p>
      <p class="chart-footnote" id="gmvSourceMissingNote" style="display:none;color:#b45309;">当前数据源中未找到 gmv source sheet，无法生成 ADG 来源占比图；请将数据补充至 Excel 后重新生成报告。</p>
    </div>

    <footer>数值轴已脱敏 · Site 名称正常显示 · 详细数据见配套 Excel 数据底表</footer>
  </div>

  <script>
    const DATA = {data_json};

    const MORANDI = [
      "#d4a574", "#9aabbd", "#a8b5a0", "#c9a9a6", "#b0a8b9",
      "#b5a898", "#8fa3a8", "#c2b280", "#a69080", "#9db4a8",
      "#b8a99a", "#c4a484", "#a3b1c2", "#b9c4aa", "#d4b896",
      "#9e8f84", "#b7a7a0"
    ];
    const ORANGE = "#d4a574";
    const LINE_COLORS = ["#d4a574", "#9aabbd", "#a8b5a0"];

    Chart.defaults.plugins.tooltip.enabled = false;
    Chart.defaults.interaction.mode = null;

    const noTooltip = {{
      enabled: false,
      external: null,
      callbacks: {{
        label: () => "",
        title: () => ""
      }}
    }};

    const chartInteraction = {{
      intersect: false,
      mode: null
    }};

    const siteAxis = {{
      ticks: {{
        display: true,
        color: "#374151",
        font: {{ size: 12, weight: "600" }}
      }},
      grid: {{ color: "#ebe7e2" }},
      title: {{ display: false }}
    }};

    const hiddenValueAxis = {{
      ticks: {{ display: false }},
      grid: {{ color: "#ebe7e2" }},
      title: {{ display: true }}
    }};

    const hiddenPercentAxis = {{
      min: 0,
      max: 100,
      ticks: {{ display: false }},
      grid: {{ color: "#ebe7e2" }},
      title: {{ display: true, text: "CR (%)" }}
    }};

    function makePie(canvasId, labels, values) {{
      new Chart(document.getElementById(canvasId), {{
        type: "pie",
        data: {{
          labels,
          datasets: [{{
            data: values,
            backgroundColor: MORANDI.slice(0, labels.length),
            borderWidth: 1,
            borderColor: "#fff"
          }}]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{
            legend: {{ position: "right" }},
            tooltip: noTooltip
          }}
        }}
      }});
    }}

    function stackedBarYMax(block) {{
      const totals = block.sites.map((_, siteIdx) =>
        block.sources.reduce(
          (sum, name) => sum + (Number(block.series[name][siteIdx]) || 0),
          0
        )
      );
      const maxTotal = Math.max(...totals, 0);
      if (maxTotal <= 0) return 100;
      return Math.ceil(maxTotal / 5) * 5;
    }}

    function makeStackedPercentBar(canvasId, block) {{
      const datasets = block.sources.map((name, idx) => ({{
        label: name,
        data: block.series[name],
        backgroundColor: MORANDI[idx % MORANDI.length],
        borderWidth: 0,
        stack: "adg_source"
      }}));
      const yMax = stackedBarYMax(block);
      return new Chart(document.getElementById(canvasId), {{
        type: "bar",
        data: {{ labels: block.sites, datasets }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{
            legend: {{ position: "bottom" }},
            tooltip: noTooltip
          }},
          scales: {{
            x: {{
              stacked: true,
              ticks: {{ display: true, color: "#374151", font: {{ size: 12, weight: "600" }} }},
              grid: {{ color: "#ebe7e2" }},
              title: {{ display: true, text: "Site" }}
            }},
            y: {{
              stacked: true,
              min: 0,
              max: yMax,
              ticks: {{ display: false }},
              grid: {{ color: "#ebe7e2" }},
              title: {{ display: true, text: "ADG 来源占比 (%)" }}
            }}
          }}
        }}
      }});
    }}

    function makeBar(canvasId, labels, values, yTitle) {{
      new Chart(document.getElementById(canvasId), {{
        type: "bar",
        data: {{
          labels,
          datasets: [{{
            data: values,
            backgroundColor: ORANGE,
            borderRadius: 4
          }}]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{ legend: {{ display: false }}, tooltip: noTooltip }},
          scales: {{
            x: siteAxis,
            y: {{ ...hiddenValueAxis, title: {{ display: true, text: yTitle }} }}
          }}
        }}
      }});
    }}

    function makeCrLine(canvasId, block) {{
      const datasets = Object.entries(block.series).map(([label, values], idx) => ({{
        label,
        data: values,
        borderColor: LINE_COLORS[idx],
        backgroundColor: LINE_COLORS[idx],
        tension: 0.2,
        pointRadius: 4,
        fill: false
      }}));
      new Chart(document.getElementById(canvasId), {{
        type: "line",
        data: {{ labels: block.sites, datasets }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{ tooltip: noTooltip }},
          scales: {{
            x: siteAxis,
            y: hiddenPercentAxis
          }}
        }}
      }});
    }}

    function renderTierMatrix(containerId, block, metricKey, title) {{
      const container = document.getElementById(containerId);
      const sites = block.sites;
      const tiers = block.tiers;
      const maxBySite = {{}};
      sites.forEach(site => {{
        maxBySite[site] = Math.max(...block[metricKey][site], 0);
      }});

      let html = `<div class="tier-matrix">`;
      html += `<div class="matrix-banner">${{title}}</div>`;
      html += `<table class="tier-grid"><thead><tr><th class="corner">Site</th>`;
      sites.forEach(site => {{ html += `<th>${{site}}</th>`; }});
      html += `</tr></thead><tbody>`;

      tiers.forEach((tier, tierIdx) => {{
        html += `<tr><td class="tier-label">${{tier}}</td>`;
        sites.forEach(site => {{
          const value = block[metricKey][site][tierIdx];
          const maxValue = maxBySite[site];
          const width = maxValue > 0 ? (value / maxValue) * 100 : 0;
          const barClass = value > 0 && value === maxValue ? "bar-max" : "bar-normal";
          html += `<td class="bar-cell"><div class="bar-track"><div class="bar ${{barClass}}" style="width:${{width}}%"></div></div></td>`;
        }});
        html += `</tr>`;
      }});

      html += `</tbody></table>`;
      html += `<div class="matrix-footnote">数据时间范围：2026年6月（单月）</div></div>`;
      container.innerHTML = html;
    }}

    makePie("ordersPie", DATA.part1.sites, DATA.part1.orders_pct);
    makePie("gmvPie", DATA.part1.sites, DATA.part1.gmv_pct);
    makeBar("aovBar", DATA.part1.sites, DATA.part1.aov, "客单价 (USD)");
    makeCrLine("crOrders", DATA.cr_orders);
    makeCrLine("crGmvSku", DATA.cr_gmv_sku);
    makeCrLine("crGmvShop", DATA.cr_gmv_shop);
    renderTierMatrix("tierGmvMatrix", DATA.price_tier, "gmv_pct", "各市场销售价格段 GMV 分布");
    renderTierMatrix("tierOrdersMatrix", DATA.price_tier, "orders_pct", "各市场销售价格段 Orders 分布");

    document.getElementById("crossSiteSummary").innerHTML =
      renderSummaryMarkdown(DATA.cross_site_summary || "");

    const TOP_SITES = Object.keys(DATA.top_skus);
    let activeSite = TOP_SITES[0];

    function escapeHtml(text) {{
      return String(text ?? "")
        .replace(/&/g, "&amp;")
        .replace(/</g, "&lt;")
        .replace(/>/g, "&gt;")
        .replace(/"/g, "&quot;");
    }}

    function renderSummaryMarkdown(text) {{
      return escapeHtml(text)
        .replace(/\\*\\*(.+?)\\*\\*/g, "<strong>$1</strong>")
        .replace(/\\n/g, "<br>");
    }}

    function renderSkuTable(containerId, items) {{
      const rows = items.map(item => `
        <tr>
          <td class="rank-col">${{item.rank}}</td>
          <td>
            <div class="name-cn">${{escapeHtml(item.translation || "-")}}</div>
            <div class="name-foreign">${{escapeHtml(item.item_name || "-")}}</div>
          </td>
          <td class="tier-col">${{escapeHtml(item.price_range || "-")}}</td>
          <td>${{
            item.item_link
              ? `<a href="${{escapeHtml(item.item_link)}}" target="_blank" rel="noopener noreferrer">打开链接</a>`
              : "-"
          }}</td>
        </tr>
      `).join("");

      document.getElementById(containerId).innerHTML = `
        <table class="sku-table">
          <thead>
            <tr>
              <th class="rank-col">Rank</th>
              <th>中文名 / 外文 SKU Name</th>
              <th class="tier-col">Price Range</th>
              <th>Item Link</th>
            </tr>
          </thead>
          <tbody>${{rows}}</tbody>
        </table>
      `;
    }}

    function renderSiteTabs() {{
      const tabs = document.getElementById("siteTabs");
      tabs.innerHTML = TOP_SITES.map(site => `
        <button class="site-tab ${{site === activeSite ? "active" : ""}}" data-site="${{site}}">${{site}}</button>
      `).join("");
      tabs.querySelectorAll(".site-tab").forEach(btn => {{
        btn.addEventListener("click", () => {{
          activeSite = btn.dataset.site;
          renderSiteView();
        }});
      }});
    }}

    function renderSiteView() {{
      const block = DATA.top_skus[activeSite];
      document.getElementById("siteSummary").innerHTML =
        renderSummaryMarkdown(DATA.summaries[activeSite] || "");
      renderSkuTable("skuOrdersTable", block.by_orders || []);
      renderSkuTable("skuGmvTable", block.by_gmv || []);
      renderSiteTabs();
    }}

    renderSiteView();

    const TREND_TABS = [
      {{ key: "orders", label: "Orders" }},
      {{ key: "orders_pct", label: "Orders Site %" }},
      {{ key: "gmv_usd", label: "GMV (USD)" }},
      {{ key: "gmv_pct", label: "GMV Site %" }},
      {{ key: "aov", label: "客单价 (USD)" }}
    ];
    let activeTrendMetric = TREND_TABS[0].key;
    let activeGrowthPeriod = "h1";
    let activeGmvGrowthPeriod = "h1";
    let trendLineChart = null;
    let trendGrowthBar = null;
    let trendGmvBandBar = null;
    let ratioLineChart = null;

    const growthBandLabelPlugin = {{
      id: "growthBandLabelPlugin",
      afterDatasetsDraw(chart) {{
        const bands = chart.config.data.bandLabels;
        if (!bands) return;
        const {{ ctx }} = chart;
        const meta = chart.getDatasetMeta(0);
        ctx.save();
        ctx.fillStyle = "#374151";
        ctx.font = "600 11px Segoe UI, PingFang SC, Microsoft YaHei, sans-serif";
        ctx.textAlign = "center";
        meta.data.forEach((bar, idx) => {{
          const label = bands[idx];
          if (!label || label === "-") return;
          const y = bar.y < bar.base ? bar.y - 6 : bar.y - 6;
          ctx.fillText(label, bar.x, y);
        }});
        ctx.restore();
      }}
    }};

    const timeAxis = {{
      ticks: {{
        display: true,
        color: "#374151",
        font: {{ size: 11, weight: "600" }},
        maxRotation: 45,
        minRotation: 0
      }},
      grid: {{ color: "#ebe7e2" }},
      title: {{ display: true, text: "时间" }}
    }};

    function makeMultiSiteLineChart(canvasId, block, yTitle, isPercent = false) {{
      const datasets = block.sites.map((site, idx) => ({{
        label: site,
        data: block.series[site],
        borderColor: MORANDI[idx % MORANDI.length],
        backgroundColor: MORANDI[idx % MORANDI.length],
        tension: 0.2,
        pointRadius: 3,
        spanGaps: false,
        fill: false
      }}));
      const yAxis = isPercent
        ? {{
            ticks: {{ display: false }},
            grid: {{ color: "#ebe7e2" }},
            title: {{ display: true, text: yTitle }}
          }}
        : {{
            ticks: {{ display: false }},
            grid: {{ color: "#ebe7e2" }},
            title: {{ display: true, text: yTitle }}
          }};
      return new Chart(document.getElementById(canvasId), {{
        type: "line",
        data: {{ labels: block.months, datasets }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{
            legend: {{ position: "bottom" }},
            tooltip: noTooltip
          }},
          scales: {{
            x: timeAxis,
            y: yAxis
          }}
        }}
      }});
    }}

    function growthToBand(v) {{
      if (v == null || Number.isNaN(v)) return "-";
      const low = Math.floor(v / 10) * 10;
      const high = low + 10;
      return `${{low}}%-${{high}}%`;
    }}

    function growthBandMidpoint(v) {{
      if (v == null || Number.isNaN(v)) return 0;
      const low = Math.floor(v / 10) * 10;
      return (low + (low + 10)) / 2;
    }}

    function makeGrowthBandBar(canvasId, sites, values) {{
      const bands = values.map(growthToBand);
      const heights = values.map(growthBandMidpoint);
      const barColors = values.map(v => (v != null && v < 0) ? "#9aabbd" : ORANGE);
      return new Chart(document.getElementById(canvasId), {{
        type: "bar",
        data: {{
          labels: sites,
          bandLabels: bands,
          datasets: [{{
            data: heights,
            backgroundColor: barColors,
            borderRadius: 4
          }}]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteraction,
          plugins: {{ legend: {{ display: false }}, tooltip: noTooltip }},
          scales: {{
            x: siteAxis,
            y: {{
              ticks: {{ display: false }},
              grid: {{ color: "#ebe7e2" }},
              title: {{ display: true, text: "增速区间（相对高低）" }}
            }}
          }}
        }},
        plugins: [growthBandLabelPlugin]
      }});
    }}

    function renderPeriodTabs(containerId, block, activeKey, onSelect) {{
      const tabs = document.getElementById(containerId);
      if (!tabs) return;
      tabs.innerHTML = block.tabs.map(tab => `
        <button class="metric-tab ${{tab.key === activeKey ? "active" : ""}}" data-period="${{tab.key}}">${{tab.label}}</button>
      `).join("");
      tabs.querySelectorAll(".metric-tab").forEach(btn => {{
        btn.addEventListener("click", () => onSelect(btn.dataset.period));
      }});
    }}

    function growthFootnoteText(metricLabel, periodTab) {{
      const periodLabel = periodTab.key === "h1"
        ? `半年增速（${{periodTab.subtitle}}）`
        : `${{periodTab.label}}（${{periodTab.subtitle}}）`;
      return `当前指标：<strong>${{metricLabel}}</strong> · ${{periodLabel}} · 柱顶为 10 个百分点增速区间；数值轴已脱敏。`;
    }}

    function renderGmvBandChart() {{
      const metricBlock = DATA.trend_growth.gmv_usd;
      const footnote = document.getElementById("trendGmvBandFootnote");
      if (!metricBlock || !metricBlock.tabs.length) {{
        if (footnote) footnote.textContent = "暂无 GMV 增速区间数据。";
        return;
      }}
      const periodTab = metricBlock.tabs.find(item => item.key === activeGmvGrowthPeriod)
        || metricBlock.tabs[0];
      activeGmvGrowthPeriod = periodTab.key;
      if (trendGmvBandBar) trendGmvBandBar.destroy();
      trendGmvBandBar = makeGrowthBandBar(
        "trendGmvBandBar",
        metricBlock.sites,
        periodTab.values
      );
      if (footnote) {{
        footnote.innerHTML = growthFootnoteText("GMV (USD)", periodTab);
      }}
      renderPeriodTabs("gmvGrowthPeriodTabs", metricBlock, activeGmvGrowthPeriod, period => {{
        activeGmvGrowthPeriod = period;
        renderGmvBandChart();
      }});
    }}

    function renderTrendMetricTabs() {{
      const tabs = document.getElementById("trendMetricTabs");
      tabs.innerHTML = TREND_TABS.map(tab => `
        <button class="metric-tab ${{tab.key === activeTrendMetric ? "active" : ""}}" data-metric="${{tab.key}}">${{tab.label}}</button>
      `).join("");
      tabs.querySelectorAll(".metric-tab").forEach(btn => {{
        btn.addEventListener("click", () => {{
          activeTrendMetric = btn.dataset.metric;
          activeGrowthPeriod = "h1";
          renderTrendChart();
        }});
      }});
    }}

    function renderTrendChart() {{
      const block = DATA.trend[activeTrendMetric];
      const tab = TREND_TABS.find(item => item.key === activeTrendMetric);
      if (trendLineChart) trendLineChart.destroy();
      trendLineChart = makeMultiSiteLineChart(
        "trendLineChart",
        block,
        tab.label,
        activeTrendMetric.endsWith("_pct")
      );
      document.getElementById("trendFootnote").innerHTML =
        `${{tab.label}} 月度趋势（每条折线为一个 Site）。数值轴已脱敏；时间轴与 Site 图例正常显示。`;
      renderTrendMetricTabs();
      renderGrowthChart();
    }}

    function renderGrowthChart() {{
      const metricBlock = DATA.trend_growth[activeTrendMetric];
      const tab = TREND_TABS.find(item => item.key === activeTrendMetric);
      if (!metricBlock || !metricBlock.tabs.length) return;
      const periodTab = metricBlock.tabs.find(item => item.key === activeGrowthPeriod)
        || metricBlock.tabs[0];
      activeGrowthPeriod = periodTab.key;
      if (trendGrowthBar) trendGrowthBar.destroy();
      trendGrowthBar = makeGrowthBandBar(
        "trendGrowthBar",
        metricBlock.sites,
        periodTab.values
      );
      document.getElementById("trendGrowthFootnote").innerHTML =
        growthFootnoteText(tab.label, periodTab);
      renderPeriodTabs("growthPeriodTabs", metricBlock, activeGrowthPeriod, period => {{
        activeGrowthPeriod = period;
        renderGrowthChart();
      }});
    }}

    function renderRatioChart() {{
      const missing = document.getElementById("ratioMissingNote");
      const footnote = document.getElementById("ratioFootnote");
      if (!DATA.has_l3_ratio || !DATA.a4_l3_ratio) {{
        missing.style.display = "block";
        footnote.style.display = "none";
        if (ratioLineChart) {{
          ratioLineChart.destroy();
          ratioLineChart = null;
        }}
        return;
      }}
      missing.style.display = "none";
      footnote.style.display = "block";
      if (ratioLineChart) ratioLineChart.destroy();
      ratioLineChart = makeMultiSiteLineChart(
        "ratioLineChart",
        DATA.a4_l3_ratio,
        "A4 GMV / L3 GMV (%)",
        true
      );
      footnote.innerHTML =
        "各 Site A4 纸 GMV 占该 Site L3（Printing & Photocopy Paper）GMV 的比例趋势。数值轴已脱敏；时间轴与 Site 图例正常显示。";
    }}

    function renderGmvSourceChart() {{
      const missing = document.getElementById("gmvSourceMissingNote");
      const footnote = document.getElementById("gmvSourceFootnote");
      if (!DATA.has_gmv_source || !DATA.gmv_source) {{
        missing.style.display = "block";
        footnote.style.display = "none";
        return;
      }}
      missing.style.display = "none";
      footnote.style.display = "block";
      makeStackedPercentBar("gmvSourceStacked", DATA.gmv_source);
      footnote.textContent = DATA.gmv_source.period_note || footnote.textContent;
    }}

    renderTrendChart();
    renderGmvBandChart();
    renderRatioChart();
    renderGmvSourceChart();
  </script>
</body>
</html>
"""
    html_path = OUTPUT_DIR / html_name
    html_path.write_text(html, encoding="utf-8")
    return html_path


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook()
    sheet_names = site_sheet_names(wb)

    ref_ws = wb[sheet_names[0]]
    part1 = read_part1(ref_ws)

    cr_orders, cr_gmv_sku, cr_gmv_shop, price_tiers, top_skus, summaries = {}, {}, {}, {}, {}, {}
    for name in sheet_names:
        ws = wb[name]
        cr_orders[name] = read_cr_block(ws, CR_ORDER_ROWS, 3)
        cr_gmv_sku[name] = read_cr_block(ws, CR_GMV_SKU_ROWS, 4)
        cr_gmv_shop[name] = read_cr_block(ws, CR_GMV_SHOP_ROWS, 7)
        price_tiers[name] = read_price_tiers(ws)
        top_skus[name] = read_top_skus(ws)
        summaries[name] = generate_site_summary(name, top_skus[name])

    cross_site_summary = generate_cross_site_summary(sheet_names, top_skus)

    y26_sheet = resolve_sheet_name(wb, Y26_SHEET_NAMES)
    l3_sheet = resolve_sheet_name(wb, L3_SHEET_NAMES)
    if not y26_sheet:
        raise ValueError(f"未找到 Y26 月度数据 sheet（尝试过: {Y26_SHEET_NAMES}）")
    y26_blocks = parse_monthly_site_blocks(wb[y26_sheet])
    l3_blocks = parse_l3_sheet(wb[l3_sheet]) if l3_sheet else []
    l3_gmv_pivot = blocks_to_gmv_pivot(l3_blocks) if l3_blocks else None
    trend_tables = build_trend_tables(y26_blocks)
    growth_tables = build_trend_growth_tables(trend_tables)
    a4_l3_ratio = build_a4_l3_ratio(y26_blocks, l3_blocks)
    if l3_sheet:
        print(f"L3 sheet found: {len(l3_blocks)} months parsed.")
    else:
        print("Warning: L3 sheet not found; A4/L3 ratio sheets and chart will be empty.")

    gmv_source = read_gmv_source(wb)
    if gmv_source:
        print(f"gmv source sheet found: {len(gmv_source['sites'])} sites parsed.")
    else:
        print("Warning: gmv source sheet not found; Part7 chart will be empty.")

    try:
        xlsx_path = build_data_workbook(
            wb,
            part1,
            cr_orders,
            cr_gmv_sku,
            cr_gmv_shop,
            price_tiers,
            top_skus,
            summaries,
            cross_site_summary,
            trend_tables,
            a4_l3_ratio,
            l3_gmv_pivot,
            growth_tables,
        )
        print(f"Excel: {xlsx_path}")
    except PermissionError:
        print("Excel: skipped (file is open, close it and re-run to refresh data workbook)")

    html_path = build_html(
        part1,
        cr_orders,
        cr_gmv_sku,
        cr_gmv_shop,
        price_tiers,
        top_skus,
        summaries,
        cross_site_summary,
        trend_tables,
        a4_l3_ratio,
        "植印A4纸_分析报告_v3.html",
        l3_gmv_pivot,
        gmv_source,
        growth_tables,
    )

    print(f"HTML:  {html_path}")


if __name__ == "__main__":
    main()
