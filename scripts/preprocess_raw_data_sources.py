"""Extract configured raw workbook sheets into AI-readable processed CSV tables."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = PROJECT_ROOT / "config" / "processed_data_sources.yaml"
DEFAULT_RAW_DIR = PROJECT_ROOT / "CateMate_rawdata"
DEFAULT_PROCESSED_DIR = PROJECT_ROOT / "CateMate_processeddata"


def main() -> None:
    parser = argparse.ArgumentParser(description="Preprocess raw workbook sheets into CateMate_processeddata.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--raw-data-dir", type=Path, default=DEFAULT_RAW_DIR)
    parser.add_argument("--processed-data-dir", type=Path, default=DEFAULT_PROCESSED_DIR)
    args = parser.parse_args()

    config = _load_yaml(args.config)
    processed_root = args.processed_data_dir
    processed_root.mkdir(parents=True, exist_ok=True)

    manifest_entries: list[dict[str, Any]] = []
    for table in config.get("tables", []):
        entry = _extract_table(table, args.raw_data_dir, processed_root, config.get("default_update_mode", "append_merge"))
        manifest_entries.append(entry)
        print(f"Extracted {entry['table_id']}: {entry['row_count']} rows -> {entry['output_csv']}")

    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "config_path": str(args.config),
        "raw_data_dir": str(args.raw_data_dir),
        "processed_data_dir": str(processed_root),
        "tables": manifest_entries,
    }
    manifest_path = processed_root / config.get("manifest_path", "processed_manifest.yaml")
    _write_yaml(manifest_path, manifest)
    print(f"Manifest: {manifest_path}")


def _extract_table(
    table: dict[str, Any],
    raw_data_dir: Path,
    processed_root: Path,
    default_update_mode: str,
) -> dict[str, Any]:
    grain = table.get("grain")
    search_dirs = _rawdata_search_dirs(raw_data_dir, grain)
    workbook_path = _find_workbook(search_dirs, table.get("source_workbook_keywords", []))
    source_sheet = table["source_sheet"]
    output_csv = processed_root / table["output_csv"]
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    update_mode = table.get("update_mode") or default_update_mode

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = _find_sheet_name(workbook.sheetnames, source_sheet)
    sheet = workbook[sheet_name]

    header, incoming_rows = _read_sheet_rows(sheet, table.get("selected_fields") or [])
    existing_count = _count_existing_rows(output_csv)
    if update_mode == "replace":
        merged_rows = incoming_rows
        added_count = len(incoming_rows)
        updated_count = 0
        retained_count = 0
    elif update_mode == "append_merge":
        merged_rows, added_count, updated_count, retained_count = _merge_existing_rows(
            output_csv=output_csv,
            header=header,
            incoming_rows=incoming_rows,
            dedupe_keys=table.get("dedupe_keys") or [],
        )
    else:
        raise ValueError(f"Unsupported update_mode for {table['table_id']}: {update_mode}")

    _write_csv(output_csv, header, merged_rows)

    source_stat = workbook_path.stat()
    return {
        "table_id": table["table_id"],
        "grain": grain or "",
        "description": table.get("description", ""),
        "source_workbook": str(workbook_path),
        "source_workbook_name": workbook_path.name,
        "source_sheet": sheet_name,
        "output_csv": str(output_csv),
        "output_csv_relative": table["output_csv"],
        "row_count": len(merged_rows),
        "incoming_row_count": len(incoming_rows),
        "existing_row_count_before": existing_count,
        "added_row_count": added_count,
        "updated_row_count": updated_count,
        "retained_existing_row_count": retained_count,
        "column_count": len(header),
        "columns": header,
        "important_fields": table.get("important_fields", []),
        "dedupe_keys": table.get("dedupe_keys") or [],
        "update_mode": update_mode,
        "source_modified_time": datetime.fromtimestamp(source_stat.st_mtime).isoformat(timespec="seconds"),
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
    }


def _read_sheet_rows(sheet: Any, selected_fields: list[str]) -> tuple[list[str], list[list[str]]]:
    header: list[str] = []
    selected_indexes: list[int] | None = None
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = _trim_trailing_empty([_normalize_cell(value) for value in row])
        if not any(value != "" for value in values):
            continue
        if not header:
            header = [str(value).strip() for value in values]
            if selected_fields:
                selected_indexes = _resolve_selected_indexes(header, selected_fields)
                header = [header[index] for index in selected_indexes]
            continue
        if selected_indexes is not None:
            values = [values[index] if index < len(values) else "" for index in selected_indexes]
        rows.append(values)
    return header, rows


def _merge_existing_rows(
    output_csv: Path,
    header: list[str],
    incoming_rows: list[list[str]],
    dedupe_keys: list[str],
) -> tuple[list[list[str]], int, int, int]:
    if not output_csv.exists():
        return incoming_rows, len(incoming_rows), 0, 0
    if not dedupe_keys:
        existing_header, existing_rows = _read_existing_csv(output_csv)
        _assert_compatible_header(header, existing_header, output_csv)
        return existing_rows + incoming_rows, len(incoming_rows), 0, len(existing_rows)

    key_indexes = list(range(len(header))) if dedupe_keys == ["__all_columns__"] else _resolve_selected_indexes(header, dedupe_keys)
    existing_header, existing_rows = _read_existing_csv(output_csv)
    _assert_compatible_header(header, existing_header, output_csv)

    merged_by_key: dict[tuple[str, ...], list[str]] = {}
    order: list[tuple[str, ...]] = []
    for row in existing_rows:
        key = _row_key(row, key_indexes)
        if key not in merged_by_key:
            order.append(key)
        merged_by_key[key] = row

    added_count = 0
    updated_count = 0
    for row in incoming_rows:
        key = _row_key(row, key_indexes)
        if key in merged_by_key:
            updated_count += 1
        else:
            added_count += 1
            order.append(key)
        merged_by_key[key] = row

    merged_rows = [merged_by_key[key] for key in order]
    retained_count = max(len(existing_rows) - updated_count, 0)
    return merged_rows, added_count, updated_count, retained_count


def _read_existing_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(header)
        writer.writerows(rows)


def _count_existing_rows(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        return max(sum(1 for _ in file) - 1, 0)


def _assert_compatible_header(expected: list[str], actual: list[str], path: Path) -> None:
    if expected != actual:
        raise ValueError(f"Existing processed table has incompatible columns: {path}")


def _row_key(row: list[str], indexes: list[int]) -> tuple[str, ...]:
    return tuple(row[index] if index < len(row) else "" for index in indexes)


def _rawdata_search_dirs(raw_data_dir: Path, grain: str | None) -> list[Path]:
    if grain:
        grain_dir = raw_data_dir / grain
        if grain_dir.is_dir():
            return [grain_dir]
    if raw_data_dir.is_dir():
        subdirs = [p for p in raw_data_dir.iterdir() if p.is_dir()]
        if subdirs:
            return sorted(subdirs)
    return [raw_data_dir]


def _find_workbook(search_dirs: list[Path], keywords: list[str]) -> Path:
    normalized_keywords = [str(keyword).lower() for keyword in keywords if str(keyword)]
    for search_dir in search_dirs:
        candidates = sorted(search_dir.glob("*.xlsx"))
        for path in candidates:
            source_name = path.name.lower()
            if all(keyword.lower() in source_name for keyword in normalized_keywords):
                return path
    raise FileNotFoundError(
        f"No workbook in {search_dirs} matched keywords: {keywords}"
    )


def _find_sheet_name(sheet_names: list[str], expected_name: str) -> str:
    for sheet_name in sheet_names:
        if sheet_name == expected_name:
            return sheet_name
    for sheet_name in sheet_names:
        if sheet_name.strip() == expected_name.strip():
            return sheet_name
    raise KeyError(f"Workbook does not contain sheet: {expected_name}")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value)


def _trim_trailing_empty(values: list[str]) -> list[str]:
    while values and values[-1] == "":
        values.pop()
    return values


def _resolve_selected_indexes(header: list[str], selected_fields: list[str]) -> list[int]:
    indexes: list[int] = []
    for field in selected_fields:
        if field not in header:
            raise ValueError(f"Selected field not found in source header: {field}")
        indexes.append(header.index(field))
    return indexes


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file) or {}
    if not isinstance(loaded, dict):
        raise ValueError(f"YAML root must be a mapping: {path}")
    return loaded


def _write_yaml(path: Path, payload: dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as file:
        yaml.safe_dump(payload, file, allow_unicode=True, sort_keys=False)


if __name__ == "__main__":
    main()
