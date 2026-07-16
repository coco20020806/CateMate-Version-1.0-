"""Build English category tree JSON from category_tree.xlsx."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE = PROJECT_ROOT / "CateMate_rawdata" / "category_tree.xlsx"
OUT_LLM_TREE = PROJECT_ROOT / "CateMate_rawdata" / "category_tree_en.json"
OUT_DIR = PROJECT_ROOT / "CateMate_processeddata"
OUT_TREE = OUT_DIR / "category_tree_en.json"
OUT_FLAT = OUT_DIR / "category_tree_en_flat.json"

HEADER = "cb_level1_global_be_category"


def is_english(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text or text == HEADER:
        return False
    ascii_ratio = sum(1 for ch in text if ord(ch) < 128) / len(text)
    return ascii_ratio > 0.8


def build_records(ws) -> list[dict]:
    current = {"l1": "", "l2": "", "l3": "", "l4": ""}
    records: list[dict] = []

    for row_idx in range(2, 281):
        l1, l2, l3, l4 = (ws.cell(row_idx, col).value for col in range(1, 5))

        if l1 == HEADER:
            continue
        if l1 and not is_english(l1):
            continue

        if l1:
            current = {"l1": str(l1).strip(), "l2": "", "l3": "", "l4": ""}
        if l2:
            current["l2"] = str(l2).strip()
            current["l3"] = ""
            current["l4"] = ""
        if l3:
            current["l3"] = str(l3).strip()
            current["l4"] = ""
        if l4:
            current["l4"] = str(l4).strip()

        if not any([l1, l2, l3, l4]):
            continue

        path_parts = [p for p in [current["l1"], current["l2"], current["l3"], current["l4"]] if p]
        records.append(
            {
                "source_row": row_idx,
                "l1": current["l1"],
                "l2": current["l2"] or None,
                "l3": current["l3"] or None,
                "l4": current["l4"] or None,
                "depth": len(path_parts),
                "path": " > ".join(path_parts),
                "path_parts": path_parts,
            }
        )

    return records


def build_nested_tree(records: list[dict]) -> dict:
    tree: dict = {}
    for record in records:
        node = tree
        for part in record["path_parts"]:
            node = node.setdefault(part, {})
    return tree


def count_nodes(node: dict) -> int:
    if not node:
        return 0
    return len(node) + sum(count_nodes(child) for child in node.values())


def build_llm_tree(nested: dict, level: int = 1) -> list[dict]:
    """Build an LLM-friendly tree where each node carries an explicit level tag."""
    tree: list[dict] = []
    for name, children in nested.items():
        node: dict = {"level": f"L{level}", "name": name}
        if children:
            node["children"] = build_llm_tree(children, level + 1)
        tree.append(node)
    return tree


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook.active
    records = build_records(worksheet)
    sheet_name = worksheet.title
    workbook.close()

    nested = build_nested_tree(records)
    l1_list = sorted(nested.keys())

    payload = {
        "meta": {
            "source_file": str(SOURCE),
            "sheet": sheet_name,
            "english_columns": ["A", "B", "C", "D"],
            "row_range": "2-280",
            "generated_on": date.today().isoformat(),
            "description": "English category tree extracted from category_tree.xlsx with hierarchical forward-fill",
            "stats": {
                "level1_count": len(l1_list),
                "record_count": len(records),
                "node_count": count_nodes(nested),
            },
        },
        "level1_categories": l1_list,
        "records": records,
        "tree": nested,
    }

    flat_payload = {
        "meta": payload["meta"],
        "paths": [record["path"] for record in records],
        "records": records,
    }

    llm_tree = build_llm_tree(nested)
    OUT_LLM_TREE.write_text(json.dumps(llm_tree, ensure_ascii=False, indent=2), encoding="utf-8")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_TREE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_FLAT.write_text(json.dumps(flat_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {OUT_LLM_TREE}")
    print(f"Wrote {OUT_TREE}")
    print(f"Wrote {OUT_FLAT}")
    print(f"Level-1 categories ({len(l1_list)}): {', '.join(l1_list)}")
    print(f"Records: {len(records)}, Nodes: {count_nodes(nested)}")


if __name__ == "__main__":
    main()
