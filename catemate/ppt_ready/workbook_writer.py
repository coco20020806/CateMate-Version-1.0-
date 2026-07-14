"""Write generic PPT-ready workbook from build result."""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook

from catemate.planning.schemas import RequirementPlanningSpec
from catemate.ppt_ready.schemas import (
    PptReadyBuildContext,
    PptReadySheetSpec,
    PptReadyWorkbookBuildResult,
)


def write_ppt_ready_workbook(
    result: PptReadyWorkbookBuildResult,
    context: PptReadyBuildContext,
    output_path,
    planning_spec: RequirementPlanningSpec,
    gate_message: str,
):
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    _write_catalog_sheet(workbook, result)
    _write_notes_sheet(workbook, result, context, planning_spec, gate_message)

    for sheet_spec in result.sheets:
        _write_chart_sheet(workbook, sheet_spec)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    result.output_path = output_path
    return output_path


def _join(values: list[str]) -> str:
    return "; ".join(v for v in values if v)


def _write_catalog_sheet(workbook: Workbook, result: PptReadyWorkbookBuildResult) -> None:
    sheet = workbook.create_sheet("ppt_data_catalog", 0)
    headers = [
        "sheet_name",
        "chart_id",
        "chart_title",
        "chart_type",
        "data_module_id",
        "source_table_ids",
        "source_workbook_names",
        "source_sheets",
        "processed_csv_paths",
        "source_rule_note",
        "missing_data_note",
        "null_reason_note",
        "output_status",
        "row_count",
        "notes",
    ]
    sheet.append(headers)
    for item in result.sheets:
        sheet.append(
            [
                item.sheet_name,
                item.chart_id,
                item.chart_title,
                item.chart_type,
                item.data_module_id,
                _join(item.source_table_ids),
                _join(item.source_workbook_names),
                _join(item.source_sheets),
                _join(item.processed_csv_paths),
                item.source_rule_note,
                item.missing_data_note,
                item.null_reason_note,
                item.output_status,
                len(item.rows),
                " | ".join(item.notes),
            ]
        )


def _write_notes_sheet(
    workbook: Workbook,
    result: PptReadyWorkbookBuildResult,
    context: PptReadyBuildContext,
    planning_spec: RequirementPlanningSpec,
    gate_message: str,
) -> None:
    sheet = workbook.create_sheet("data_notes", 1)
    sheet.append(["item", "value"])
    rows: list[tuple[str, Any]] = [
        ("case_id", result.case_id or context.case_id),
        ("generated_at", datetime.now(timezone.utc).replace(microsecond=0).isoformat()),
        ("planning_spec_path", str(context.planning_spec_path)),
        ("requirement_workbook_path", str(context.requirement_workbook_path)),
        ("processed_manifest_path", str(context.processed_manifest_path)),
        ("confirmation_gate_status", gate_message),
    ]

    for table_id, lineage in sorted(result.used_table_lineage.items()):
        columns = lineage.get("columns") or []
        if isinstance(columns, list) and len(columns) > 12:
            columns_text = ",".join(str(c) for c in columns[:12]) + ",..."
        elif isinstance(columns, list):
            columns_text = ",".join(str(c) for c in columns)
        else:
            columns_text = str(columns)
        if lineage.get("error"):
            value = f"error={lineage['error']}"
        else:
            value = (
                f"processed_table={table_id}; "
                f"source_workbook={lineage.get('source_workbook_name', '')}; "
                f"source_sheet={lineage.get('source_sheet', '')}; "
                f"processed_csv={lineage.get('processed_csv_path', '')}; "
                f"row_count={lineage.get('row_count', '')}; "
                f"columns={columns_text}"
            )
        rows.append((f"source_table::{table_id}", value))

    for item in result.sheets:
        chart_value = (
            f"sheet={item.sheet_name}; "
            f"chart_type={item.chart_type}; "
            f"data_module={item.data_module_id}; "
            f"source_tables={_join(item.source_table_ids)}; "
            f"source_workbook={_join(item.source_workbook_names)}; "
            f"source_sheet={_join(item.source_sheets)}"
        )
        rows.append((f"chart_source::{item.chart_id}", chart_value))
        if item.missing_data_note:
            rows.append((f"missing_reason::{item.chart_id}", item.missing_data_note))
        if item.null_reason_note:
            rows.append((f"null_reason::{item.chart_id}", item.null_reason_note))

    for index, assumption in enumerate(planning_spec.assumptions, start=1):
        rows.append((f"planning_assumption_{index}", assumption))
    for index, note in enumerate(planning_spec.source_notes, start=1):
        rows.append((f"planning_source_note_{index}", note))
    for index, warning in enumerate(result.warnings, start=1):
        rows.append((f"warning_{index}", warning))
    if not planning_spec.assumptions:
        rows.append(("planning_assumption", ""))
    if not planning_spec.source_notes:
        rows.append(("planning_source_note", ""))
    if not result.warnings:
        rows.append(("warning", ""))

    for item, value in rows:
        sheet.append([item, value])


def _builder_notes_text(spec: PptReadySheetSpec) -> str:
    parts = list(spec.notes)
    if spec.missing_data_note:
        parts.append(f"missing_data_note: {spec.missing_data_note}")
    if spec.null_reason_note:
        parts.append(f"null_reason_note: {spec.null_reason_note}")
    if spec.source_rule_note:
        parts.append(f"source_rule_note: {spec.source_rule_note}")
    return " | ".join(parts)


def _write_chart_sheet(workbook: Workbook, sheet_spec: PptReadySheetSpec) -> None:
    sheet = workbook.create_sheet(sheet_spec.sheet_name)
    rows = sheet_spec.rows
    note_text = _builder_notes_text(sheet_spec)
    if not rows:
        sheet.append(["note", "builder_notes"])
        sheet.append(["No data rows.", note_text or None])
        return

    headers: list[str] = list(rows[0].keys())
    for row in rows[1:]:
        for key in row.keys():
            if key not in headers:
                headers.append(str(key))
    if "builder_notes" not in headers:
        headers.append("builder_notes")

    sheet.append(headers)
    for index, row in enumerate(rows):
        values = [row.get(header) for header in headers if header != "builder_notes"]
        values.append(note_text if index == 0 else None)
        sheet.append(values)
