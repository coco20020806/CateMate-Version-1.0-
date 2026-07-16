"""Build HTML preview from V2 Data Workbook sheets."""

from __future__ import annotations

import html
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def read_data_workbook_tables(workbook_path: Path) -> dict[str, pd.DataFrame]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    tables: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Data."):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        data = [list(r) for r in rows[1:]]
        table_id = sheet_name.replace("Data.", "", 1)
        tables[table_id] = pd.DataFrame(data, columns=header)
    return tables


def build_html_from_data_workbook(workbook_path: Path) -> Path:
    tables = read_data_workbook_tables(workbook_path)
    if not tables:
        raise ValueError(f"No Data.* sheets found in {workbook_path}")

    sections: list[str] = []
    for table_id, df in tables.items():
        headers = "".join(f"<th>{html.escape(str(c))}</th>" for c in df.columns)
        body_rows = []
        for row in df.head(100).itertuples(index=False):
            cells = "".join(f"<td>{html.escape(str(v))}</td>" for v in row)
            body_rows.append(f"<tr>{cells}</tr>")
        sections.append(
            f"<h2>{html.escape(table_id)}</h2>"
            f"<table><thead><tr>{headers}</tr></thead>"
            f"<tbody>{''.join(body_rows)}</tbody></table>"
        )

    output_path = workbook_path.with_name(workbook_path.stem + "_preview.html")
    output_path.write_text(
        "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Data Workbook Preview</title>"
        "<style>table{border-collapse:collapse}td,th{border:1px solid #ccc;padding:4px 8px}</style>"
        "</head><body>"
        + "".join(sections)
        + "</body></html>",
        encoding="utf-8",
    )
    return output_path
