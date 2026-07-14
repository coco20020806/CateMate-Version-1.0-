"""Write updated confirmation statuses back into a requirement workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from catemate.core.confirmation_gate import ALLOWED_FINAL_STATUSES
from catemate.core.confirmation_reader import (
    CONFIRMATION_SHEET_NAME,
    HEADER_BLOCK,
    HEADER_STATUS,
)


def save_confirmation_updates(
    source_path: Path,
    status_by_row: dict[int, str],
    output_path: Path,
) -> Path:
    """Copy the workbook and update confirmation statuses / block flags.

    ``status_by_row`` maps a 1-based sheet row index (as returned by
    ``read_confirmation_records``) to its new status. The "是否阻止" column is
    recomputed from the gate rules so the two columns stay consistent.
    """
    workbook = load_workbook(source_path)
    if CONFIRMATION_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain sheet: {CONFIRMATION_SHEET_NAME}")

    sheet = workbook[CONFIRMATION_SHEET_NAME]
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    if HEADER_STATUS not in header:
        raise ValueError(f"Confirmation sheet is missing header: {HEADER_STATUS}")

    status_col = header.index(HEADER_STATUS) + 1
    block_col = header.index(HEADER_BLOCK) + 1 if HEADER_BLOCK in header else None

    for row_index, status in status_by_row.items():
        sheet.cell(row=row_index, column=status_col).value = status
        if block_col is not None:
            sheet.cell(row=row_index, column=block_col).value = "\u5426" if status in {item.value for item in ALLOWED_FINAL_STATUSES} else "\u662f"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
