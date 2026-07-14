"""Read confirmation records from a requirement workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from catemate.core.confirmation_gate import ConfirmationItem


CONFIRMATION_SHEET_NAME = "\u786e\u8ba4\u8bb0\u5f55"
HEADER_NAME = "\u786e\u8ba4\u9879"
HEADER_SUGGESTED_VALUE = "\u5efa\u8bae\u503c"
HEADER_STATUS = "\u72b6\u6001"
HEADER_REASON = "\u539f\u56e0"
HEADER_BLOCK = "\u662f\u5426\u963b\u6b62PPT-ready\u751f\u6210"
HEADER_SOURCE = "\u6765\u6e90"
HEADER_PLANNING_QUESTION_ID = "\u89c4\u5212\u95ee\u9898ID"

# Summary rows appended at the bottom of the confirmation sheet; not real items.
SUMMARY_ROW_NAMES = {
    "\u5141\u8bb8\u751f\u6210PPT-ready workbook\u7684\u72b6\u6001",
    "\u963b\u585e\u72b6\u6001",
}


def read_confirmation_items(workbook_path: Path) -> list[ConfirmationItem]:
    """Read confirmation items from the workbook's confirmation sheet."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if CONFIRMATION_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain sheet: {CONFIRMATION_SHEET_NAME}")

    sheet = workbook[CONFIRMATION_SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    indexes = _build_header_indexes(header)
    items: list[ConfirmationItem] = []

    for row in rows[1:]:
        name = _cell(row, indexes[HEADER_NAME])
        status = _cell(row, indexes[HEADER_STATUS])
        if not name or not status:
            continue
        if name in SUMMARY_ROW_NAMES:
            continue
        items.append(
            ConfirmationItem(
                name=name,
                suggested_value=_cell(row, indexes.get(HEADER_SUGGESTED_VALUE)),
                status=status,
                reason=_cell(row, indexes.get(HEADER_REASON)),
                blocks_ppt_ready=_parse_block_flag(_cell(row, indexes.get(HEADER_BLOCK))),
                source=_cell(row, indexes.get(HEADER_SOURCE)),
                planning_question_id=_cell(row, indexes.get(HEADER_PLANNING_QUESTION_ID)),
            )
        )
    return items


def read_confirmation_records(workbook_path: Path) -> list[dict[str, object]]:
    """Read confirmation rows as plain dicts, keeping their 1-based sheet row index.

    The ``row`` key lets a writer update the exact cell later, which is safer than
    matching by name because some items (e.g. 类目映射) can repeat.
    """
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if CONFIRMATION_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain sheet: {CONFIRMATION_SHEET_NAME}")

    sheet = workbook[CONFIRMATION_SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    indexes = _build_header_indexes(header)
    records: list[dict[str, object]] = []

    for offset, row in enumerate(rows[1:], start=2):
        name = _cell(row, indexes[HEADER_NAME])
        status = _cell(row, indexes[HEADER_STATUS])
        if not name or not status:
            continue
        if name in SUMMARY_ROW_NAMES:
            continue
        records.append(
            {
                "row": offset,
                HEADER_NAME: name,
                HEADER_SUGGESTED_VALUE: _cell(row, indexes.get(HEADER_SUGGESTED_VALUE)),
                HEADER_STATUS: status,
                HEADER_REASON: _cell(row, indexes.get(HEADER_REASON)),
                HEADER_BLOCK: _cell(row, indexes.get(HEADER_BLOCK)),
                HEADER_SOURCE: _cell(row, indexes.get(HEADER_SOURCE)),
                HEADER_PLANNING_QUESTION_ID: _cell(row, indexes.get(HEADER_PLANNING_QUESTION_ID)),
            }
        )
    return records


def _build_header_indexes(header: list[str]) -> dict[str, int]:
    required = [HEADER_NAME, HEADER_STATUS]
    indexes = {name: header.index(name) for name in header if name}
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"Confirmation sheet is missing required headers: {', '.join(missing)}")
    return indexes


def _cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return str(value).strip() if value is not None else ""


def _parse_block_flag(value: str) -> bool | None:
    normalized = value.strip().lower()
    if normalized in {"\u662f", "yes", "true", "1", "y"}:
        return True
    if normalized in {"\u5426", "no", "false", "0", "n"}:
        return False
    return None
