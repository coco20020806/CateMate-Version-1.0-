"""Load tables and metadata from V2 Data Workbook."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from catemate.html_report.schemas import ChartBinding


TIME_FIELDS = {"grass_month", "month", "year_month", "grass_date", "date"}
SITE_FIELDS = {"grass_region", "region", "site"}
METRIC_FIELDS = {
    "orders",
    "gmv_usd",
    "gmv",
    "aov",
    "rank",
    "item_name",
    "item_link",
    "shop_id",
    "share",
}
HEADER_HINTS = TIME_FIELDS | SITE_FIELDS | METRIC_FIELDS | {
    "item_id",
    "item_price_usd",
    "orders_pct",
    "gmv_usd_pct",
    "aov_pct",
    "subset_orders_share",
    "orders_share",
}


def parse_table_id_from_sheet(sheet_name: str) -> str:
    """Extract table_id from Data.{section}.{table_id} or Data.{run}.{table_id}."""
    if not sheet_name.startswith("Data."):
        return sheet_name
    parts = sheet_name.split(".", 2)
    if len(parts) >= 3:
        return parts[2]
    remainder = sheet_name.replace("Data.", "", 1)
    if "." in remainder:
        return remainder.split(".", 1)[1]
    return remainder


def parse_run_or_section_from_sheet(sheet_name: str) -> str:
    if not sheet_name.startswith("Data."):
        return ""
    parts = sheet_name.split(".", 2)
    if len(parts) >= 3:
        return parts[1]
    return ""


def parse_section_id_from_sheet(sheet_name: str) -> str:
    return parse_run_or_section_from_sheet(sheet_name)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_metadata_row(row: tuple[object, ...]) -> bool:
    cells = [_normalize_header(c) for c in row]
    non_empty = [c for c in cells if c]
    if not non_empty:
        return True
    if len(non_empty) == 1 and (non_empty[0].startswith("scope:") or non_empty[0].startswith("注")):
        return True
    return False


def _is_header_row(row: tuple[object, ...]) -> bool:
    cells = [_normalize_header(c).lower() for c in row]
    non_empty = [c for c in cells if c]
    if len(non_empty) < 2:
        return False
    return bool(set(non_empty) & HEADER_HINTS)


def find_header_row_index(rows: list[tuple[object, ...]]) -> int:
    for index, row in enumerate(rows):
        if _is_metadata_row(row):
            continue
        if _is_header_row(row):
            return index
    return 0


def _dedupe_column_names(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for raw in columns:
        name = _normalize_header(raw) or "column"
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}_{count + 1}")
    return result


def dataframe_from_sheet_rows(rows: list[tuple[object, ...]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    header_index = find_header_row_index(rows)
    header = _dedupe_column_names([_normalize_header(c) for c in rows[header_index]])
    data_rows = rows[header_index + 1 :]
    records = [list(row) for row in data_rows if any(cell is not None and str(cell).strip() != "" for cell in row)]
    if not records:
        return pd.DataFrame(columns=header)
    width = len(header)
    normalized_records = [list(record[:width]) + [None] * max(0, width - len(record)) for record in records]
    return pd.DataFrame(normalized_records, columns=header)


@dataclass
class WorkbookTableEntry:
    sheet_name: str
    table_id: str
    run_or_section: str
    df: pd.DataFrame


def table_id_matches(sheet_table_id: str, plan_table_id: str) -> bool:
    if sheet_table_id == plan_table_id:
        return True
    trimmed = sheet_table_id.rstrip("_")
    if trimmed and plan_table_id.startswith(trimmed):
        return True
    if plan_table_id.rstrip("_") == trimmed:
        return True
    return False


def load_workbook_table_entries(workbook_path: Path) -> list[WorkbookTableEntry]:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    entries: list[WorkbookTableEntry] = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Data."):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        df = dataframe_from_sheet_rows(rows)
        entries.append(
            WorkbookTableEntry(
                sheet_name=sheet_name,
                table_id=parse_table_id_from_sheet(sheet_name),
                run_or_section=parse_run_or_section_from_sheet(sheet_name),
                df=df,
            )
        )
    if not entries:
        raise ValueError(f"No Data.* sheets found in {workbook_path}")
    return entries


def load_workbook_tables(workbook_path: Path) -> dict[str, pd.DataFrame]:
    """Return sheet_name -> DataFrame mapping with metadata rows stripped."""
    return {entry.sheet_name: entry.df for entry in load_workbook_table_entries(workbook_path)}


def load_workbook_tables_by_sheet_table_id(workbook_path: Path) -> dict[str, pd.DataFrame]:
    """Return parsed trailing table_id -> DataFrame (last sheet wins on collision)."""
    tables: dict[str, pd.DataFrame] = {}
    for entry in load_workbook_table_entries(workbook_path):
        tables[entry.table_id] = entry.df
    return tables


def resolve_table_for_binding(
    entries: list[WorkbookTableEntry],
    *,
    table_id: str,
    run_id: str = "",
    section_id: str = "",
    sheet_name: str = "",
) -> WorkbookTableEntry | None:
    if sheet_name:
        for entry in entries:
            if entry.sheet_name == sheet_name:
                return entry

    if run_id:
        for entry in entries:
            if entry.run_or_section == run_id and table_id_matches(entry.table_id, table_id):
                return entry

    if section_id:
        for entry in entries:
            if entry.run_or_section == section_id and table_id_matches(entry.table_id, table_id):
                return entry

    for entry in entries:
        if table_id_matches(entry.table_id, table_id):
            return entry
    return None


def _pick_time_field(columns: list[str]) -> str | None:
    for name in columns:
        if name.strip().lower() in TIME_FIELDS:
            return name
    return None


def _pick_site_field(columns: list[str]) -> str | None:
    for name in columns:
        if name.strip().lower() in SITE_FIELDS:
            return name
    return None


def _pick_metric_fields(columns: list[str], chart_type: str) -> list[str]:
    if chart_type == "trend" and any("_mom" in c.lower() for c in columns):
        mom = find_mom_y_fields(columns)
        if mom:
            return mom[:1]
    if chart_type == "share":
        share = find_share_y_fields(columns)
        if share:
            return share[:1]
    metrics = [
        c
        for c in columns
        if c.strip().lower() not in TIME_FIELDS | SITE_FIELDS
        and not c.startswith("scope:")
        and not c.startswith("注")
    ]
    preferred = [c for c in metrics if c.strip().lower() in METRIC_FIELDS]
    if preferred:
        return preferred[:3]
    numeric_like = [c for c in metrics if c and not c.endswith("_1")][:3]
    return numeric_like


def repair_chart_binding(binding: ChartBinding, df: pd.DataFrame) -> ChartBinding:
    columns = [str(c) for c in df.columns if str(c).strip()]
    if not columns:
        return binding.model_copy(update={"visible": False, "confidence": "low"})

    x_field = binding.x_field if binding.x_field in columns else None
    y_fields = [field for field in binding.y_fields if field in columns]
    series_field = binding.series_field if binding.series_field in columns else None

    if binding.chart_type in {"trend", "bar", "share"}:
        if x_field is None:
            x_field = _pick_time_field(columns) if binding.chart_type == "trend" else _pick_site_field(columns)
            if x_field is None and binding.chart_type != "trend":
                x_field = columns[0]
        if not y_fields:
            y_fields = _pick_metric_fields(columns, binding.chart_type)
        if binding.chart_type == "trend" and series_field is None:
            series_field = _pick_site_field(columns)

    if binding.chart_type == "table" and not y_fields:
        y_fields = columns[: min(8, len(columns))]

    if binding.chart_type in {"trend", "bar", "share"} and (x_field is None or not y_fields):
        return binding.model_copy(
            update={
                "x_field": x_field,
                "y_fields": y_fields,
                "series_field": series_field,
                "visible": False,
                "confidence": "low",
            }
        )

    return binding.model_copy(
        update={
            "x_field": x_field,
            "y_fields": y_fields,
            "series_field": series_field,
        }
    )


def infer_table_role(table_id: str) -> tuple[str, str]:
    """Return (chart_type_hint, role) from table_id naming convention."""
    lower = table_id.lower()
    if lower.endswith("_by_site_month") or lower.endswith("_by_site_mont"):
        return "trend", "primary"
    if (
        lower.endswith("_latest_month_by_site")
        or lower.endswith("_latest_month_by_")
        or lower.endswith("_latest_month_by_sit")
        or lower.endswith("_latest_site")
        or "_latest_" in lower
    ):
        return "bar", "secondary"
    if lower.endswith("_mom_by_site_month") or lower.endswith("_mom_by_site_mont"):
        return "trend", "secondary"
    if lower.endswith("_latest_month_pct_by_site") or "_latest_month_pct" in lower or lower.endswith("_pct_by_site"):
        return "share", "secondary"
    if "share" in lower:
        return "share", "primary"
    if "rank" in lower or "top_" in lower or "listing" in lower:
        return "table", "primary"
    return "table", "primary"


def find_mom_y_fields(columns: list[str]) -> list[str]:
    return [c for c in columns if re.search(r"_mom_pct$|_mom$|mom_pct", c, re.I)]


def find_share_y_fields(columns: list[str]) -> list[str]:
    return [c for c in columns if re.search(r"_pct$|share|proportion", c, re.I)]
