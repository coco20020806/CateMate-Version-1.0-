"""First MVP module: build a category analysis data requirement workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from catemate.data.category_tree import build_sph_category_tree_lookup, find_category_candidates
from catemate.data.source_scanner import check_raw_data_fields, check_required_sheets, scan_excel_sources
from catemate.planning.requirement_adapter import build_requirement_spec_from_planning
from catemate.planning.schemas import RequirementPlanningSpec
from catemate.schemas.category_requirement import (
    AnalysisPlanRow,
    CategoryAnalysisCaseConfig,
    CategoryAnalysisRequirementSpec,
    CategoryCandidateRow,
    ChartDataRequirementRow,
    ConfirmationTemplateItem,
    DataRequirementRow,
    PreprocessPlanRow,
    RequirementContext,
    RequirementSummaryRow,
    SourceCheckRow,
    SourceFile,
)
from catemate.schemas.confirmation import ConfirmationItem
from catemate.schemas.enums import ConfirmationStatus


REQUIRED_SHEETS = ["SPH类目树", "Raw data", "CNCB 中间表 By site"]
REQUIRED_RAW_FIELDS = [
    "grass_region",
    "cb_level1_global_be_category",
    "level2_global_be_category",
    "level3_global_be_category",
    "grass_month",
    "gmv_usd",
    "orders",
]
DEMO_CATEGORY_KEYWORDS = ["Collectible", "Action Figures", "Action Figurines", "Movies & Anime", "Anime"]
MAX_CATEGORY_CONFIRMATION_ITEMS = 8


def build_requirement_workbook(
    context: RequirementContext,
    raw_data_dir: Path,
    processed_data_dir: Path,
    output_path: Path,
    case_config: CategoryAnalysisCaseConfig | None = None,
    planning_spec: RequirementPlanningSpec | None = None,
) -> Path:
    """Create the first category analysis data requirement workbook draft."""
    spec = build_category_analysis_requirement_spec(
        context=context,
        raw_data_dir=raw_data_dir,
        processed_data_dir=processed_data_dir,
        case_config=case_config,
        planning_spec=planning_spec,
    )
    return write_category_analysis_requirement_workbook(spec, output_path)


def build_requirement_workbook_from_planning(
    case_config: CategoryAnalysisCaseConfig,
    planning_spec: RequirementPlanningSpec,
    raw_data_dir: Path,
    processed_data_dir: Path,
    output_path: Path,
) -> Path:
    """Build a requirement workbook by merging case config with a planning spec."""
    context = RequirementContext(
        original_request=case_config.original_request,
        target_category_text=case_config.target_category_text,
        business_background=case_config.business_background,
        delivery_audience=case_config.delivery_audience,
        delivery_format=case_config.delivery_format,
        target_sites=case_config.target_sites,
        time_range=case_config.time_range,
    )
    return build_requirement_workbook(
        context=context,
        raw_data_dir=raw_data_dir,
        processed_data_dir=processed_data_dir,
        output_path=output_path,
        case_config=case_config,
        planning_spec=planning_spec,
    )


def build_category_analysis_requirement_spec(
    context: RequirementContext,
    raw_data_dir: Path,
    processed_data_dir: Path,
    case_config: CategoryAnalysisCaseConfig | None = None,
    planning_spec: RequirementPlanningSpec | None = None,
) -> CategoryAnalysisRequirementSpec:
    """Build a structured requirement spec without writing Excel."""
    source_files = scan_excel_sources(raw_data_dir)
    primary_source = _select_primary_source(source_files, case_config)
    category_tree_source = _select_category_tree_source(source_files, case_config)

    sheet_checks = []
    field_checks = []
    candidates: list[dict[str, str]] = []
    category_lookup_path = processed_data_dir / "sph_category_tree_lookup.csv"

    if primary_source:
        sheet_checks = check_required_sheets(primary_source.path, _resolve_required_sheets(case_config))
        field_checks = check_raw_data_fields(
            primary_source.path,
            _resolve_required_fields(case_config),
            sheet_name=_resolve_raw_fields_sheet_name(case_config),
        )

    if category_tree_source:
        tree_sheet_checks = check_required_sheets(category_tree_source.path, [_resolve_category_tree_sheet_name(case_config)])
        if primary_source and category_tree_source.path != primary_source.path:
            sheet_checks.extend(tree_sheet_checks)
        if any(check.sheet_name == _resolve_category_tree_sheet_name(case_config) and check.exists for check in tree_sheet_checks):
            lookup_rows = build_sph_category_tree_lookup(category_tree_source.path, category_lookup_path)
            candidates = find_category_candidates(lookup_rows, _resolve_category_keywords(case_config))

    has_source = primary_source is not None
    resolved_context = _resolve_context(context, case_config)
    project_name = case_config.project_name if case_config else "HKCB Collectible Category Insight 样例"
    base_spec = CategoryAnalysisRequirementSpec(
        project_name=project_name,
        requirement_summary=_build_requirement_summary(resolved_context, primary_source, project_name, category_tree_source),
        category_candidates=_build_category_candidates(candidates),
        analysis_plan=_resolve_analysis_plan(case_config),
        data_requirements=_resolve_data_requirements(case_config),
        source_checks=_build_source_checks(source_files, sheet_checks, field_checks),
        preprocess_plan=_resolve_preprocess_plan(case_config, category_lookup_path),
        chart_requirements=_resolve_chart_requirements(case_config),
        confirmation_items=_build_confirmation_items(candidates, has_source, case_config),
    )
    if planning_spec is None:
        return base_spec

    case_payload = case_config.model_dump(by_alias=True) if case_config is not None else {}
    return build_requirement_spec_from_planning(
        case_config=case_payload,
        planning_spec=planning_spec,
        base_spec=base_spec,
    )


def _select_primary_source(
    source_files: list[SourceFile],
    case_config: CategoryAnalysisCaseConfig | None,
) -> SourceFile | None:
    if not source_files:
        return None
    if case_config and case_config.source_file_keywords:
        keywords = [item.lower() for item in case_config.source_file_keywords if item]
        for source in source_files:
            source_name = source.path.name.lower()
            if all(keyword in source_name for keyword in keywords):
                return source
    return next(
        (item for item in source_files if item.matched_source_id),
        source_files[0],
    )


def _select_category_tree_source(
    source_files: list[SourceFile],
    case_config: CategoryAnalysisCaseConfig | None,
) -> SourceFile | None:
    if not source_files:
        return None
    if case_config and case_config.category_tree_source_keywords:
        return _select_source_by_keywords(source_files, case_config.category_tree_source_keywords)
    sheet_name = _resolve_category_tree_sheet_name(case_config)
    for source in source_files:
        try:
            checks = check_required_sheets(source.path, [sheet_name])
        except Exception:
            continue
        if checks and checks[0].exists:
            return source
    return None


def _select_source_by_keywords(source_files: list[SourceFile], keywords: list[str]) -> SourceFile | None:
    normalized_keywords = [item.lower() for item in keywords if item]
    for source in source_files:
        source_name = source.path.name.lower()
        if all(keyword in source_name for keyword in normalized_keywords):
            return source
    return None


def write_category_analysis_requirement_workbook(
    spec: CategoryAnalysisRequirementSpec,
    output_path: Path,
) -> Path:
    """Write a CategoryAnalysisRequirementSpec to an Excel workbook."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_requirement_summary(workbook, spec)
    _write_category_candidates(workbook, spec.category_candidates)
    _write_analysis_plan(workbook, spec.analysis_plan)
    _write_data_requirements(workbook, spec.data_requirements)
    _write_source_check(workbook, spec.source_checks)
    _write_preprocess_plan(workbook, spec.preprocess_plan)
    _write_chart_ppt_requirements(workbook, spec.chart_requirements)
    _write_confirmation_log(workbook, spec)

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _build_requirement_summary(
    context: RequirementContext,
    primary_source: SourceFile | None,
    project_name: str,
    category_tree_source: SourceFile | None = None,
) -> list[RequirementSummaryRow]:
    source_label = str(primary_source.path) if primary_source else "未找到"
    category_tree_label = str(category_tree_source.path) if category_tree_source else "未找到"
    rows = [
        ("项目", project_name),
        ("原始需求", context.original_request),
        ("目标类目文本", context.target_category_text),
        ("业务背景", context.business_background),
        ("交付对象", context.delivery_audience),
        ("期望形式", context.delivery_format),
        ("时间范围", context.time_range),
        ("源数据文件", source_label),
        ("类目树文件", category_tree_label),
        ("当前状态", "数据需求 workbook 草稿，待用户确认类目映射和缺失数据"),
    ]
    return [RequirementSummaryRow(field=field, content=content) for field, content in rows]


def _build_category_candidates(candidates: list[dict[str, str]]) -> list[CategoryCandidateRow]:
    if not candidates:
        return [
            CategoryCandidateRow(
                confirmation_status="\u5f85\u8865\u5145",
                note="\u672a\u627e\u5230\u5019\u9009\u7c7b\u76ee\u3002",
            )
        ]
    return [
        CategoryCandidateRow(
            user_text=item.get("matched_text", ""),
            candidate_path=item["category_path"],
            l1=item["l1"],
            l2=item["l2"],
            l3=item["l3"],
            match_type=item.get("match_type", ""),
            confirmation_status="\u5f85\u786e\u8ba4",
            note="\u524d\u53f0\u7c7b\u76ee\u4e0e\u540e\u53f0\u7c7b\u76ee\u53ef\u80fd\u4e0d\u5b8c\u5168\u540c\u540d\uff0c\u9700\u8981\u7528\u6237\u786e\u8ba4\u3002",
        )
        for item in candidates
    ]


def _default_analysis_plan() -> list[AnalysisPlanRow]:
    rows = [
        ("类目总览", "目标类目在各站点的 GMV、Orders、客单价与同比", "支持", "Raw data", "按站点、月份、类目聚合"),
        ("站点分布", "各站点销售占比与贡献", "支持", "Raw data", "可用于 PPT 分布图"),
        ("L3 分布", "L2 下各 L3 的规模与增长", "支持", "Raw data + 类目树", "需要确认目标 L2/L3"),
        ("YoY 气泡图数据", "规模、增速、客单价三维展示", "支持", "Raw data", "参考 CNCB 中间表 By site 公式"),
        ("价格段分布", "各价格带订单量和 GMV 贡献", "暂不支持", "需要 SKU/商品价格明细", "当前源数据无价格字段"),
        ("关键词搜索量同比", "Related keywords Search Volume YOY", "暂不支持", "需要关键词搜索量数据源", "当前源数据无关键词字段"),
    ]
    return [
        AnalysisPlanRow(
            analysis_block=block,
            question=question,
            support_status=status,
            dependencies=deps,
            note=note,
        )
        for block, question, status, deps, note in rows
    ]


def _default_data_requirements() -> list[DataRequirementRow]:
    rows = [
        ("SPH 月度类目数据", "Raw data", "是", "底层 GMV/Orders 聚合", "无法生成趋势和站点分布", "已在样例文件中找到"),
        ("SPH 月度类目数据", "SPH类目树", "是", "类目映射和层级查找", "无法判断 L1/L2/L3 对应关系", "已预处理为查找表"),
        ("SPH 月度类目数据", "CNCB 中间表 By site", "建议", "理解现有公式和气泡图结构", "可自行复刻，但缺少人工公式参考", "已在样例文件中找到"),
        ("价格明细数据", "price / price_band / item_id", "否", "价格段分析", "不能输出价格段分布", "当前缺失"),
        ("关键词数据", "keyword / search_volume / month", "否", "搜索量同比", "不能输出关键词搜索量 YoY", "当前缺失"),
    ]
    return [
        DataRequirementRow(
            data_source=source,
            field_or_sheet=field_or_sheet,
            is_required=is_required,
            purpose=purpose,
            missing_impact=missing_impact,
            current_note=current_note,
        )
        for source, field_or_sheet, is_required, purpose, missing_impact, current_note in rows
    ]


def _build_source_checks(source_files, sheet_checks, field_checks) -> list[SourceCheckRow]:
    rows: list[SourceCheckRow] = []
    for source in source_files:
        rows.append(
            SourceCheckRow(
                check_type="文件",
                object_name=source.path.name,
                status="已找到",
                note=f"{source.size_bytes} bytes; modified {source.modified_time}",
            )
        )
    if not source_files:
        rows.append(
            SourceCheckRow(
                check_type="文件",
                object_name="CateMate_rawdata",
                status="未找到",
                note="请放入 SPH Excel 源数据",
            )
        )
    for check in sheet_checks:
        rows.append(
            SourceCheckRow(
                check_type="Sheet",
                object_name=check.sheet_name,
                status="已找到" if check.exists else "缺失",
                note=check.note,
            )
        )
    for check in field_checks:
        rows.append(
            SourceCheckRow(
                check_type="字段",
                object_name=check.field_name,
                status="已找到" if check.exists else "缺失",
                note=check.note,
            )
        )
    return rows


def _default_preprocess_plan(lookup_path: Path) -> list[PreprocessPlanRow]:
    rows = [
        ("类目树扁平化", "SPH类目树", str(lookup_path), "向下继承空白层级，形成 L1/L2/L3/L4 查找表"),
        ("Raw data 标准化", "Raw data", "待实现", "统一月份、站点、类目字段和数值类型"),
        ("分析块聚合", "标准化 Raw data", "待实现", "按站点、类目、月份聚合 GMV 和 Orders"),
    ]
    return [
        PreprocessPlanRow(step=step, input_name=input_name, output_name=output_name, note=note)
        for step, input_name, output_name, note in rows
    ]


def _default_chart_requirements() -> list[ChartDataRequirementRow]:
    rows = [
        ("L1 总览", "category_overview_by_site", "site, gmv, orders, yoy, abs", "可生成", "用于总览概括"),
        ("L2 站点分布", "site_distribution", "site, gmv_share, order_share", "可生成", "用于销售占比分布图"),
        ("L3 品类分布", "l3_distribution", "l3, site, gmv, orders, yoy", "可生成", "需要确认目标 L2"),
        ("YoY 气泡图", "yoy_bubble_chart", "site_or_l3, gmv, yoy, abs", "可生成", "参考中间表公式"),
        ("价格段分布", "price_band_distribution", "price_band, gmv, orders, contribution", "缺数据", "需要价格明细"),
        ("关键词搜索量", "keyword_search_yoy", "keyword, month, search_volume, yoy", "缺数据", "需要关键词数据源"),
    ]
    return [
        ChartDataRequirementRow(
            chart_page=chart_page,
            required_table=required_table,
            fields=fields,
            status=status,
            note=note,
        )
        for chart_page, required_table, fields, status, note in rows
    ]


def _build_confirmation_items(
    candidates: list[dict[str, str]],
    has_source: bool,
    case_config: CategoryAnalysisCaseConfig | None = None,
) -> list[ConfirmationItem]:
    template_items = _resolve_confirmation_templates(case_config)
    items: list[ConfirmationItem] = []
    for template in template_items:
        status = template.status
        if template.name == "源数据文件":
            status = ConfirmationStatus.CONFIRMED if has_source else ConfirmationStatus.PENDING_SUPPLEMENT
        question_text = template.question.strip() or template.reason.strip()
        items.append(
            ConfirmationItem(
                name=template.name,
                suggested_value=template.suggested_value,
                status=status,
                reason=question_text,
            )
        )

    for item in candidates[:MAX_CATEGORY_CONFIRMATION_ITEMS]:
        items.append(
            ConfirmationItem(
                name="类目映射",
                suggested_value=item["category_path"],
                status=ConfirmationStatus.PENDING_CONFIRMATION,
                reason=f"由 {item.get('matched_text', '')} 匹配得到",
            )
        )
    return items


def _resolve_context(
    context: RequirementContext,
    case_config: CategoryAnalysisCaseConfig | None,
) -> RequirementContext:
    if not case_config:
        return context
    return RequirementContext(
        original_request=case_config.original_request,
        target_category_text=case_config.target_category_text,
        business_background=case_config.business_background,
        delivery_audience=case_config.delivery_audience,
        delivery_format=case_config.delivery_format,
        target_sites=case_config.target_sites,
        time_range=case_config.time_range,
    )


def _resolve_category_keywords(case_config: CategoryAnalysisCaseConfig | None) -> list[str]:
    if case_config and case_config.category_keywords:
        return case_config.category_keywords
    return DEMO_CATEGORY_KEYWORDS


def _resolve_required_sheets(case_config: CategoryAnalysisCaseConfig | None) -> list[str]:
    if case_config and case_config.required_sheets:
        return case_config.required_sheets
    return REQUIRED_SHEETS


def _resolve_raw_fields_sheet_name(case_config: CategoryAnalysisCaseConfig | None) -> str:
    if case_config and case_config.raw_fields_sheet_name:
        return case_config.raw_fields_sheet_name
    return "Raw data"


def _resolve_required_fields(case_config: CategoryAnalysisCaseConfig | None) -> list[str]:
    if case_config and case_config.required_fields:
        return case_config.required_fields
    return REQUIRED_RAW_FIELDS


def _resolve_category_tree_sheet_name(case_config: CategoryAnalysisCaseConfig | None) -> str:
    if case_config and case_config.category_tree_sheet_name:
        return case_config.category_tree_sheet_name
    return "SPH类目树"


def _resolve_analysis_plan(case_config: CategoryAnalysisCaseConfig | None) -> list[AnalysisPlanRow]:
    if case_config and case_config.analysis_plan:
        return case_config.analysis_plan
    return _default_analysis_plan()


def _resolve_data_requirements(case_config: CategoryAnalysisCaseConfig | None) -> list[DataRequirementRow]:
    if case_config and case_config.data_requirements:
        return case_config.data_requirements
    return _default_data_requirements()


def _resolve_preprocess_plan(
    case_config: CategoryAnalysisCaseConfig | None,
    category_lookup_path: Path,
) -> list[PreprocessPlanRow]:
    if not case_config or not case_config.preprocess_plan:
        return _default_preprocess_plan(category_lookup_path)

    resolved_rows: list[PreprocessPlanRow] = []
    for row in case_config.preprocess_plan:
        output_name = row.output_name.replace("{category_lookup_path}", str(category_lookup_path))
        resolved_rows.append(
            PreprocessPlanRow(
                step=row.step,
                input_name=row.input_name,
                output_name=output_name,
                note=row.note,
            )
        )
    return resolved_rows


def _resolve_chart_requirements(case_config: CategoryAnalysisCaseConfig | None) -> list[ChartDataRequirementRow]:
    if case_config and case_config.chart_requirements:
        return case_config.chart_requirements
    return _default_chart_requirements()


def _resolve_confirmation_templates(
    case_config: CategoryAnalysisCaseConfig | None,
) -> list[ConfirmationTemplateItem]:
    if case_config and case_config.confirmation_templates:
        return case_config.confirmation_templates
    return [
        ConfirmationTemplateItem(
            name="源数据文件",
            suggested_value="使用当前 SPH RM Excel",
            status=ConfirmationStatus.PENDING_SUPPLEMENT,
            reason="按文件名匹配",
        ),
        ConfirmationTemplateItem(
            name="交付敏感性",
            suggested_value="对外交付，需脱敏",
            status=ConfirmationStatus.PENDING_CONFIRMATION,
            reason="招商/卖家沟通场景",
        ),
        ConfirmationTemplateItem(
            name="时间范围",
            suggested_value="使用源数据最新完整月份",
            status=ConfirmationStatus.PENDING_CONFIRMATION,
            reason="用户未指定明确时间",
        ),
        ConfirmationTemplateItem(
            name="价格段分析",
            suggested_value="需要额外价格明细数据",
            status=ConfirmationStatus.PENDING_CONFIRMATION,
            reason="当前源数据不支持",
        ),
        ConfirmationTemplateItem(
            name="关键词搜索量同比",
            suggested_value="需要额外关键词数据",
            status=ConfirmationStatus.PENDING_CONFIRMATION,
            reason="当前源数据不支持",
        ),
    ]


def _write_requirement_summary(workbook: Workbook, spec: CategoryAnalysisRequirementSpec) -> None:
    sheet = workbook.create_sheet("需求摘要")
    sheet.append(["字段", "内容"])
    for row in spec.requirement_summary:
        sheet.append([row.field, row.content])


def _write_category_candidates(workbook: Workbook, candidates: list[CategoryCandidateRow]) -> None:
    sheet = workbook.create_sheet("类目映射候选")
    sheet.append(["用户文本", "候选后台路径", "L1", "L2", "L3", "匹配方式", "确认状态", "说明"])
    for item in candidates:
        sheet.append(
            [
                item.user_text,
                item.candidate_path,
                item.l1,
                item.l2,
                item.l3,
                item.match_type,
                item.confirmation_status,
                item.note,
            ]
        )


def _write_analysis_plan(workbook: Workbook, rows: list[AnalysisPlanRow]) -> None:
    sheet = workbook.create_sheet("分析计划")
    has_planning = any(row.module_id or row.planning_reason for row in rows)
    headers = ["分析块", "回答问题", "当前支持状态", "依赖数据", "备注"]
    if has_planning:
        headers.extend(["数据模块ID", "规划理由"])
    sheet.append(headers)
    for row in rows:
        values = [row.analysis_block, row.question, row.support_status, row.dependencies, row.note]
        if has_planning:
            values.extend([row.module_id, row.planning_reason])
        sheet.append(values)


def _write_data_requirements(workbook: Workbook, rows: list[DataRequirementRow]) -> None:
    sheet = workbook.create_sheet("数据需求清单")
    has_planning = any(row.module_id or row.table_id or row.planning_reason or row.source_notes for row in rows)
    headers = ["数据源", "字段/Sheet", "是否必须", "用途", "缺失影响", "当前备注"]
    if has_planning:
        headers.extend(["数据模块ID", "数据表ID", "规划理由", "来源说明"])
    sheet.append(headers)
    for row in rows:
        values = [
            row.data_source,
            row.field_or_sheet,
            row.is_required,
            row.purpose,
            row.missing_impact,
            row.current_note,
        ]
        if has_planning:
            values.extend([row.module_id, row.table_id, row.planning_reason, row.source_notes])
        sheet.append(values)


def _write_source_check(workbook: Workbook, rows: list[SourceCheckRow]) -> None:
    sheet = workbook.create_sheet("源数据检查")
    sheet.append(["检查类型", "对象", "状态", "说明"])
    for row in rows:
        sheet.append([row.check_type, row.object_name, row.status, row.note])


def _write_preprocess_plan(workbook: Workbook, rows: list[PreprocessPlanRow]) -> None:
    sheet = workbook.create_sheet("预处理规划")
    sheet.append(["步骤", "输入", "输出", "说明"])
    for row in rows:
        sheet.append([row.step, row.input_name, row.output_name, row.note])


def _write_chart_ppt_requirements(workbook: Workbook, rows: list[ChartDataRequirementRow]) -> None:
    sheet = workbook.create_sheet("图表PPT数据需求")
    has_planning = any(
        row.data_module_id or row.table_ids or row.grain or row.metrics or row.dimensions or row.planning_reason
        for row in rows
    )
    has_extended = any(
        row.chart_intent or row.x_axis or row.y_axis or row.series or row.sort_rule or row.optional_flag or row.selection_reason
        for row in rows
    )
    headers = ["图表/页面", "所需数据表", "字段", "当前状态", "备注"]
    if has_planning:
        headers.extend(["数据模块ID", "数据表ID列表", "数据粒度", "指标", "维度", "规划理由"])
    if has_extended:
        headers.extend(["图表意图", "X轴", "Y轴", "系列", "排序规则", "是否可选", "模块选择理由"])
    sheet.append(headers)
    for row in rows:
        values = [row.chart_page, row.required_table, row.fields, row.status, row.note]
        if has_planning:
            values.extend(
                [
                    row.data_module_id,
                    row.table_ids,
                    row.grain,
                    row.metrics,
                    row.dimensions,
                    row.planning_reason,
                ]
            )
        if has_extended:
            values.extend(
                [
                    row.chart_intent,
                    row.x_axis,
                    row.y_axis,
                    row.series,
                    row.sort_rule,
                    row.optional_flag,
                    row.selection_reason,
                ]
            )
        sheet.append(values)


def _write_confirmation_log(workbook: Workbook, spec: CategoryAnalysisRequirementSpec) -> None:
    sheet = workbook.create_sheet("确认记录")
    has_planning = any(item.source or item.planning_question_id for item in spec.confirmation_items)
    headers = ["确认项", "建议值", "状态", "原因", "是否阻止PPT-ready生成"]
    if has_planning:
        headers.extend(["来源", "规划问题ID"])
    sheet.append(headers)
    allowed_status_values = {status.value for status in spec.allowed_final_statuses}

    for item in spec.confirmation_items:
        status_value = item.status.value if isinstance(item.status, ConfirmationStatus) else str(item.status)
        if item.blocks_ppt_ready is True:
            block_value = "是"
        elif item.blocks_ppt_ready is False:
            block_value = "否"
        else:
            block_value = "否" if status_value in allowed_status_values else "是"
        values = [item.name, item.suggested_value, status_value, item.reason, block_value]
        if has_planning:
            values.extend([item.source, item.planning_question_id])
        sheet.append(values)

    sheet.append([])
    empty_trail = [""] * (2 if has_planning else 0)
    sheet.append(
        [
            "允许生成PPT-ready workbook的状态",
            "、".join(sorted(status.value for status in spec.allowed_final_statuses)),
            "",
            "其他状态都会阻止生成",
            "",
            *empty_trail,
        ]
    )
    sheet.append(
        [
            "阻塞状态",
            "、".join(sorted(status.value for status in spec.blocking_statuses)),
            "",
            "用户补充数据后需先复检，不能直接生成",
            "",
            *empty_trail,
        ]
    )


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8)
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)
    sheet.freeze_panes = "A2"
