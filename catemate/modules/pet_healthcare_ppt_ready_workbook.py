"""Generate PPT-ready tables for the VN Pet Healthcare validation case."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from catemate.schemas import CategoryLevel, ChartType, DataSourceStatus, PptReadyTableSpec, PptReadyWorkbookSpec


SOURCE_SHEETS = {
    "history": "\u8fc7\u53bb\u6570\u636e",
    "price_tier": "price tier",
    "top_listing": "\u70ed\u95e8\u5546\u54c1",
}
TARGET_SITE = "VN"
TARGET_L1 = "Pets"
TARGET_L2 = "Pet Healthcare"
CONFIRMED_MAPPING_SCOPE = f"{TARGET_L1} > {TARGET_L2}"


class PetHealthcareContext(BaseModel):
    source_workbook_path: Path
    output_path: Path


def build_pet_healthcare_ppt_ready_workbook(context: PetHealthcareContext) -> Path:
    """Build PPT-ready tables for VN Pet Healthcare."""
    history = _load_history(context.source_workbook_path)
    price_tier = _load_price_tier(context.source_workbook_path)
    top_listing = _load_top_listing(context.source_workbook_path)

    target_history = _filter_target(history)
    target_price_tier = _filter_target(price_tier)
    target_top_listing = _filter_target(top_listing)

    latest_month = target_history["grass_month"].max()
    prior_year_month = latest_month - pd.DateOffset(years=1)
    latest_price_month = target_price_tier["year_month"].max() if not target_price_tier.empty else pd.NaT
    workbook_spec = _workbook_spec()

    sheets = {
        "ppt_data_catalog": _catalog(workbook_spec, latest_month, prior_year_month, latest_price_month),
        "data_notes": _data_notes(context, workbook_spec, latest_month, prior_year_month, latest_price_month),
        "vn_pet_health_trend": _market_trend(target_history, context.source_workbook_path),
        "vn_pet_health_price_tier": _price_tier_distribution(target_price_tier, latest_price_month, context.source_workbook_path),
        "vn_pet_health_avg_price": _average_price_by_l3(target_top_listing, context.source_workbook_path),
        "vn_pet_health_top_listing": _top_listing(target_top_listing, context.source_workbook_path),
    }

    context.output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(context.output_path, engine="openpyxl") as writer:
        for sheet_name, data_frame in sheets.items():
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_workbook(writer.book)
    return context.output_path


def _workbook_spec() -> PptReadyWorkbookSpec:
    return PptReadyWorkbookSpec(
        workbook_name="vn_pet_healthcare_ppt_ready",
        confirmed_mapping=CONFIRMED_MAPPING_SCOPE,
        target_site=TARGET_SITE,
        target_category=TARGET_L2,
        tables=[
            PptReadyTableSpec(
                sheet_name="vn_pet_health_trend",
                chart_types=[ChartType.TREND, ChartType.BAR],
                grain="VN Pet Healthcare by month",
                category_level=CategoryLevel.L2,
                source_sheets=[SOURCE_SHEETS["history"]],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="GMV, orders, ABS, and YoY helper fields.",
            ),
            PptReadyTableSpec(
                sheet_name="vn_pet_health_price_tier",
                chart_types=[ChartType.SHARE, ChartType.BAR],
                grain="VN Pet Healthcare by price tier",
                category_level=CategoryLevel.L2,
                source_sheets=[SOURCE_SHEETS["price_tier"]],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="ADG/ADO distribution by price range.",
            ),
            PptReadyTableSpec(
                sheet_name="vn_pet_health_top_listing",
                chart_types=[ChartType.TABLE],
                grain="VN Pet Healthcare item listing",
                category_level=CategoryLevel.L3,
                source_sheets=[SOURCE_SHEETS["top_listing"]],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="Top listing data sorted by current ADGMV.",
            ),
            PptReadyTableSpec(
                sheet_name="vn_pet_health_avg_price",
                chart_types=[ChartType.BAR, ChartType.TABLE],
                grain="VN Pet Healthcare top listings by L3",
                category_level=CategoryLevel.L3,
                source_sheets=[SOURCE_SHEETS["top_listing"]],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="Average and median listing price from top-listing source, not full SKU universe.",
            ),
        ],
    )


def _load_history(path: Path) -> pd.DataFrame:
    data = pd.read_excel(path, sheet_name=SOURCE_SHEETS["history"], engine="openpyxl")
    fields = [
        "grass_region",
        "cb_level1_global_be_category",
        "level2_global_be_category",
        "level3_global_be_category",
        "grass_month",
        "gmv_usd",
        "orders",
    ]
    _require_fields(data, fields, SOURCE_SHEETS["history"])
    data = data[fields].copy()
    data["grass_month"] = pd.to_datetime(data["grass_month"])
    data["gmv_usd"] = pd.to_numeric(data["gmv_usd"], errors="coerce").fillna(0)
    data["orders"] = pd.to_numeric(data["orders"], errors="coerce").fillna(0)
    return data


def _load_price_tier(path: Path) -> pd.DataFrame:
    data = pd.read_excel(path, sheet_name=SOURCE_SHEETS["price_tier"], engine="openpyxl")
    fields = [
        "year_month",
        "grass_region",
        "cb_level1_global_be_category",
        "level2_global_be_category",
        "level3_global_be_category",
        "Price_Range_USD",
        "ADO",
        "ADG",
        "Live_SKUs",
    ]
    _require_fields(data, fields, SOURCE_SHEETS["price_tier"])
    data = data[fields].copy()
    data["year_month"] = pd.to_datetime(data["year_month"])
    for field in ["ADO", "ADG", "Live_SKUs"]:
        data[field] = pd.to_numeric(data[field], errors="coerce").fillna(0)
    return data


def _load_top_listing(path: Path) -> pd.DataFrame:
    data = pd.read_excel(path, sheet_name=SOURCE_SHEETS["top_listing"], engine="openpyxl")
    fields = [
        "item_name",
        "item_link",
        "item_image",
        "grass_region",
        "cb_level1_global_be_category",
        "level2_global_be_category",
        "level3_global_be_category",
        "item_price_usd",
        "current_ado(RAW)",
        "current_adgmv(RAW)",
    ]
    _require_fields(data, fields, SOURCE_SHEETS["top_listing"])
    data = data[fields].copy()
    for field in ["item_price_usd", "current_ado(RAW)", "current_adgmv(RAW)"]:
        data[field] = pd.to_numeric(data[field], errors="coerce").fillna(0)
    return data


def _filter_target(data: pd.DataFrame) -> pd.DataFrame:
    return data[
        (data["grass_region"] == TARGET_SITE)
        & (data["cb_level1_global_be_category"] == TARGET_L1)
        & (data["level2_global_be_category"] == TARGET_L2)
    ].copy()


def _catalog(
    workbook_spec: PptReadyWorkbookSpec,
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
    latest_price_month: pd.Timestamp,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "sheet_name": table.sheet_name,
                "chart_type": " / ".join(chart.value for chart in table.chart_types),
                "grain": table.grain,
                "status": table.status.value,
                "notes": table.calculation_note,
            }
            for table in workbook_spec.tables
        ]
    ).assign(
        metric_period=latest_month.strftime("%Y-%m") if pd.notna(latest_month) else "",
        comparison_period=prior_year_month.strftime("%Y-%m") if pd.notna(prior_year_month) else "",
        price_tier_period=latest_price_month.strftime("%Y-%m") if pd.notna(latest_price_month) else "",
    )


def _data_notes(
    context: PetHealthcareContext,
    workbook_spec: PptReadyWorkbookSpec,
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
    latest_price_month: pd.Timestamp,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"field": "source_file", "value": str(context.source_workbook_path)},
            {"field": "target_site", "value": TARGET_SITE},
            {"field": "target_l1", "value": TARGET_L1},
            {"field": "target_l2", "value": TARGET_L2},
            {"field": "confirmed_mapping", "value": workbook_spec.confirmed_mapping},
            {"field": "trend_source_sheet", "value": SOURCE_SHEETS["history"]},
            {"field": "price_tier_source_sheet", "value": SOURCE_SHEETS["price_tier"]},
            {"field": "top_listing_source_sheet", "value": SOURCE_SHEETS["top_listing"]},
            {"field": "metric_period", "value": latest_month.strftime("%Y-%m") if pd.notna(latest_month) else ""},
            {"field": "comparison_period", "value": prior_year_month.strftime("%Y-%m") if pd.notna(prior_year_month) else ""},
            {"field": "price_tier_period", "value": latest_price_month.strftime("%Y-%m") if pd.notna(latest_price_month) else ""},
            {"field": "yoy_policy", "value": "latest month vs same month last year for market trend helper fields"},
        ]
    )


def _market_trend(data: pd.DataFrame, source_path: Path) -> pd.DataFrame:
    grouped = data.groupby("grass_month", as_index=False).agg(
        gmv_usd=("gmv_usd", "sum"),
        orders=("orders", "sum"),
    )
    grouped["abs"] = (grouped["gmv_usd"] / grouped["orders"].where(grouped["orders"] != 0)).fillna(0)
    grouped = grouped.sort_values("grass_month")
    grouped["gmv_usd_prior_year"] = grouped["gmv_usd"].shift(12).fillna(0)
    grouped["orders_prior_year"] = grouped["orders"].shift(12).fillna(0)
    grouped["gmv_yoy"] = _safe_growth(grouped["gmv_usd"], grouped["gmv_usd_prior_year"])
    grouped["orders_yoy"] = _safe_growth(grouped["orders"], grouped["orders_prior_year"])
    grouped.insert(0, "source_file", source_path.name)
    grouped.insert(1, "source_sheet", SOURCE_SHEETS["history"])
    grouped.insert(2, "category_level", "L2")
    grouped.insert(3, "category_name", TARGET_L2)
    grouped.insert(4, "parent_category", TARGET_L1)
    grouped.insert(5, "site", TARGET_SITE)
    grouped["metric_period"] = grouped["grass_month"].dt.strftime("%Y-%m")
    grouped = grouped.drop(columns=["grass_month"])
    grouped["calculation_note"] = "Monthly VN Pet Healthcare trend from history sheet."
    grouped["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return grouped


def _price_tier_distribution(data: pd.DataFrame, latest_month: pd.Timestamp, source_path: Path) -> pd.DataFrame:
    if pd.isna(latest_month):
        return pd.DataFrame()
    current = data[data["year_month"] == latest_month].copy()
    grouped = current.groupby("Price_Range_USD", as_index=False).agg(
        ado=("ADO", "sum"),
        adg=("ADG", "sum"),
        live_skus=("Live_SKUs", "sum"),
    )
    grouped["ado_share"] = grouped["ado"] / grouped["ado"].sum() if grouped["ado"].sum() else 0
    grouped["adg_share"] = grouped["adg"] / grouped["adg"].sum() if grouped["adg"].sum() else 0
    grouped.insert(0, "source_file", source_path.name)
    grouped.insert(1, "source_sheet", SOURCE_SHEETS["price_tier"])
    grouped.insert(2, "category_level", "L2")
    grouped.insert(3, "category_name", TARGET_L2)
    grouped.insert(4, "parent_category", TARGET_L1)
    grouped.insert(5, "site", TARGET_SITE)
    grouped["metric_period"] = latest_month.strftime("%Y-%m")
    grouped["calculation_note"] = "Latest available price-tier data for VN Pet Healthcare."
    grouped["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return grouped.sort_values("Price_Range_USD")


def _top_listing(data: pd.DataFrame, source_path: Path, limit: int = 50) -> pd.DataFrame:
    ranked = data.sort_values("current_adgmv(RAW)", ascending=False).head(limit).copy()
    ranked = ranked.rename(
        columns={
            "level3_global_be_category": "category_name",
            "current_ado(RAW)": "current_ado",
            "current_adgmv(RAW)": "current_adgmv",
        }
    )
    ranked = ranked[
        [
            "category_name",
            "item_name",
            "item_link",
            "item_image",
            "item_price_usd",
            "current_ado",
            "current_adgmv",
        ]
    ].copy()
    ranked.insert(0, "source_file", source_path.name)
    ranked.insert(1, "source_sheet", SOURCE_SHEETS["top_listing"])
    ranked.insert(2, "category_level", "L3")
    ranked.insert(4, "parent_category", TARGET_L2)
    ranked.insert(5, "site", TARGET_SITE)
    ranked["metric_period"] = "current"
    ranked["calculation_note"] = "Top listing sorted by current_adgmv."
    ranked["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return ranked


def _average_price_by_l3(data: pd.DataFrame, source_path: Path) -> pd.DataFrame:
    grouped = data.groupby("level3_global_be_category", as_index=False).agg(
        listing_count=("item_name", "count"),
        avg_item_price_usd=("item_price_usd", "mean"),
        median_item_price_usd=("item_price_usd", "median"),
        total_current_ado=("current_ado(RAW)", "sum"),
        total_current_adgmv=("current_adgmv(RAW)", "sum"),
    )
    grouped = grouped.rename(columns={"level3_global_be_category": "category_name"})
    grouped.insert(0, "source_file", source_path.name)
    grouped.insert(1, "source_sheet", SOURCE_SHEETS["top_listing"])
    grouped.insert(2, "category_level", "L3")
    grouped.insert(4, "parent_category", TARGET_L2)
    grouped.insert(5, "site", TARGET_SITE)
    grouped["metric_period"] = "current"
    grouped["calculation_note"] = "Average price from top-listing source; not full SKU universe."
    grouped["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return grouped.sort_values("total_current_adgmv", ascending=False)


def _safe_growth(current: pd.Series, prior: pd.Series) -> pd.Series:
    return ((current / prior.where(prior != 0)) - 1).fillna(0)


def _require_fields(data: pd.DataFrame, fields: list[str], sheet_name: str) -> None:
    missing = [field for field in fields if field not in data.columns]
    if missing:
        raise ValueError(f"{sheet_name} is missing required fields: {', '.join(missing)}")


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
        sheet.freeze_panes = "A2"
