"""Generate a minimal PPT-ready workbook after confirmation gate passes."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from catemate.core.confirmation_gate import STATUS_CONFIRMED, STATUS_NOT_NEEDED, evaluate_confirmation_gate
from catemate.core.confirmation_reader import read_confirmation_items
from catemate.schemas import CategoryLevel, ChartType, DataSourceStatus, PptReadyTableSpec, PptReadyWorkbookSpec


RAW_SHEET_NAME = "Raw data"
L1_CATEGORY = "Hobbies & Collections"
L2_CATEGORY = "Collectible Items"
CONFIRMED_MAPPING_SCOPE = f"{L1_CATEGORY} > {L2_CATEGORY}"


class PptReadyContext(BaseModel):
    confirmed_workbook_path: Path
    raw_workbook_path: Path
    output_path: Path


def build_ppt_ready_workbook(context: PptReadyContext) -> Path:
    """Build the first PPT-ready workbook from confirmed inputs and SPH raw data."""
    confirmation_items = read_confirmation_items(context.confirmed_workbook_path)
    gate_result = evaluate_confirmation_gate(confirmation_items)
    if not gate_result.can_generate:
        raise ValueError("Confirmation gate has not passed; cannot generate PPT-ready workbook.")

    selected_l3 = _selected_l3_categories(confirmation_items)
    raw_data = _load_raw_data(context.raw_workbook_path)
    target_data = raw_data[
        (raw_data["cb_level1_global_be_category"] == L1_CATEGORY)
        & (raw_data["level2_global_be_category"] == L2_CATEGORY)
    ].copy()

    latest_month = target_data["grass_month"].max()
    prior_year_month = latest_month - pd.DateOffset(years=1)
    workbook_spec = _workbook_spec(selected_l3)

    sheets = {
        "ppt_data_catalog": _ppt_data_catalog(workbook_spec, latest_month, prior_year_month),
        "data_notes": _data_notes(context, workbook_spec, selected_l3, latest_month, prior_year_month),
        "site_performance_l2": _site_performance_l2(target_data, latest_month, prior_year_month, context.raw_workbook_path),
        "l3_distribution": _l3_distribution(target_data, latest_month, prior_year_month, context.raw_workbook_path),
        "monthly_trend_by_site": _monthly_trend_by_site(target_data, context.raw_workbook_path),
    }

    context.output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(context.output_path, engine="openpyxl") as writer:
        for sheet_name, data_frame in sheets.items():
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_workbook(writer.book)

    return context.output_path


def _selected_l3_categories(confirmation_items) -> list[str]:
    selected: list[str] = []
    for item in confirmation_items:
        if item.name == "类目映射" and item.status == STATUS_CONFIRMED:
            candidate = item.suggested_value.split(" > ")[-1].strip()
            if candidate and candidate not in selected:
                selected.append(candidate)
    return selected


def _workbook_spec(selected_l3: list[str]) -> PptReadyWorkbookSpec:
    return PptReadyWorkbookSpec(
        workbook_name="collectible_items_ppt_ready",
        confirmed_mapping=CONFIRMED_MAPPING_SCOPE,
        target_category=L2_CATEGORY,
        tables=[
            PptReadyTableSpec(
                sheet_name="site_performance_l2",
                chart_types=[ChartType.BAR, ChartType.BUBBLE],
                grain="L2 by site",
                category_level=CategoryLevel.L2,
                source_sheets=[RAW_SHEET_NAME],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="GMV, orders, ABS, and YoY by site.",
            ),
            PptReadyTableSpec(
                sheet_name="l3_distribution",
                chart_types=[ChartType.SHARE, ChartType.BAR],
                grain="L2 by L3",
                category_level=CategoryLevel.L3,
                source_sheets=[RAW_SHEET_NAME],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="L3 GMV/orders distribution and YoY.",
            ),
            PptReadyTableSpec(
                sheet_name="monthly_trend_by_site",
                chart_types=[ChartType.TREND],
                grain="L2 by site by month",
                category_level=CategoryLevel.L2,
                source_sheets=[RAW_SHEET_NAME],
                status=DataSourceStatus.AVAILABLE,
                calculation_note="Monthly GMV, orders, and ABS by site.",
            ),
            PptReadyTableSpec(
                sheet_name="selected_l3_by_site",
                chart_types=[ChartType.BAR, ChartType.TREND],
                grain="selected L3 by site",
                category_level=CategoryLevel.L3,
                status=DataSourceStatus.PARTIAL,
                calculation_note=f"Planned next. Selected L3: {', '.join(selected_l3) if selected_l3 else 'none'}.",
            ),
            PptReadyTableSpec(
                sheet_name="yoy_bubble_data",
                chart_types=[ChartType.BUBBLE],
                grain="site or L3",
                category_level=CategoryLevel.UNKNOWN,
                status=DataSourceStatus.PARTIAL,
                calculation_note="Planned next after first workbook structure is accepted.",
            ),
        ],
        notes=["HTML preview is reserved, not generated in v1."],
    )


def _load_raw_data(raw_workbook_path: Path) -> pd.DataFrame:
    data = pd.read_excel(raw_workbook_path, sheet_name=RAW_SHEET_NAME, engine="openpyxl")
    required = [
        "grass_region",
        "cb_level1_global_be_category",
        "level2_global_be_category",
        "level3_global_be_category",
        "grass_month",
        "gmv_usd",
        "orders",
    ]
    missing = [field for field in required if field not in data.columns]
    if missing:
        raise ValueError(f"Raw data is missing required fields: {', '.join(missing)}")

    data = data[required].copy()
    data["grass_month"] = pd.to_datetime(data["grass_month"])
    data["gmv_usd"] = pd.to_numeric(data["gmv_usd"], errors="coerce").fillna(0)
    data["orders"] = pd.to_numeric(data["orders"], errors="coerce").fillna(0)
    return data


def _ppt_data_catalog(
    workbook_spec: PptReadyWorkbookSpec,
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
) -> pd.DataFrame:
    catalog = pd.DataFrame(
        [
            {
                "sheet_name": table.sheet_name,
                "chart_type": " / ".join(chart.value for chart in table.chart_types),
                "grain": table.grain,
                "status": table.status.value,
                "notes": table.calculation_note,
            }
            for table in workbook_spec.tables
        ]
    )
    catalog["metric_period"] = latest_month.strftime("%Y-%m")
    catalog["comparison_period"] = prior_year_month.strftime("%Y-%m")
    return catalog


def _data_notes(
    context: PptReadyContext,
    workbook_spec: PptReadyWorkbookSpec,
    selected_l3: list[str],
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"field": "confirmed_workbook", "value": str(context.confirmed_workbook_path)},
            {"field": "source_file", "value": str(context.raw_workbook_path)},
            {"field": "source_sheet", "value": RAW_SHEET_NAME},
            {"field": "target_l1", "value": L1_CATEGORY},
            {"field": "target_l2", "value": L2_CATEGORY},
            {"field": "confirmed_mapping", "value": workbook_spec.confirmed_mapping},
            {"field": "selected_l3", "value": ", ".join(selected_l3) if selected_l3 else ""},
            {"field": "metric_period", "value": latest_month.strftime("%Y-%m")},
            {"field": "comparison_period", "value": prior_year_month.strftime("%Y-%m")},
            {"field": "yoy_policy", "value": "latest selected month vs same month last year, following current intermediate-table logic"},
            {"field": "currency", "value": "USD"},
            {"field": "desensitization", "value": "not applied at PPT-ready workbook layer"},
            {"field": "preview_html", "value": "reserved, not generated in v1"},
        ]
    )


def _site_performance_l2(
    data: pd.DataFrame,
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
    raw_workbook_path: Path,
) -> pd.DataFrame:
    current = _aggregate(data[data["grass_month"] == latest_month], ["grass_region"])
    prior = _aggregate(data[data["grass_month"] == prior_year_month], ["grass_region"])
    merged = _merge_yoy(current, prior, ["grass_region"])
    merged.insert(0, "source_file", raw_workbook_path.name)
    merged.insert(1, "source_sheet", RAW_SHEET_NAME)
    merged.insert(2, "category_level", "L2")
    merged.insert(3, "category_name", L2_CATEGORY)
    merged.insert(4, "parent_category", L1_CATEGORY)
    merged = merged.rename(columns={"grass_region": "site"})
    merged["metric_period"] = latest_month.strftime("%Y-%m")
    merged["calculation_note"] = "Aggregated by site; YoY compares latest month with same month last year."
    merged["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return merged


def _l3_distribution(
    data: pd.DataFrame,
    latest_month: pd.Timestamp,
    prior_year_month: pd.Timestamp,
    raw_workbook_path: Path,
) -> pd.DataFrame:
    current = _aggregate(data[data["grass_month"] == latest_month], ["level3_global_be_category"])
    prior = _aggregate(data[data["grass_month"] == prior_year_month], ["level3_global_be_category"])
    merged = _merge_yoy(current, prior, ["level3_global_be_category"])
    total_gmv = merged["gmv_usd"].sum()
    total_orders = merged["orders"].sum()
    merged["gmv_share"] = merged["gmv_usd"] / total_gmv if total_gmv else 0
    merged["orders_share"] = merged["orders"] / total_orders if total_orders else 0
    merged.insert(0, "source_file", raw_workbook_path.name)
    merged.insert(1, "source_sheet", RAW_SHEET_NAME)
    merged.insert(2, "category_level", "L3")
    merged = merged.rename(columns={"level3_global_be_category": "category_name"})
    merged.insert(4, "parent_category", L2_CATEGORY)
    merged.insert(5, "site", "ALL")
    merged["metric_period"] = latest_month.strftime("%Y-%m")
    merged["calculation_note"] = "Aggregated by L3 across all sites; share uses latest month."
    merged["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return merged.sort_values("gmv_usd", ascending=False)


def _monthly_trend_by_site(data: pd.DataFrame, raw_workbook_path: Path) -> pd.DataFrame:
    trend = _aggregate(data, ["grass_month", "grass_region"])
    trend.insert(0, "source_file", raw_workbook_path.name)
    trend.insert(1, "source_sheet", RAW_SHEET_NAME)
    trend.insert(2, "category_level", "L2")
    trend.insert(3, "category_name", L2_CATEGORY)
    trend.insert(4, "parent_category", L1_CATEGORY)
    trend = trend.rename(columns={"grass_region": "site"})
    trend["metric_period"] = trend["grass_month"].dt.strftime("%Y-%m")
    trend = trend.drop(columns=["grass_month"])
    trend["calculation_note"] = "Monthly aggregated by site."
    trend["confirmed_mapping"] = CONFIRMED_MAPPING_SCOPE
    return trend.sort_values(["metric_period", "site"])


def _aggregate(data: pd.DataFrame, group_fields: list[str]) -> pd.DataFrame:
    grouped = data.groupby(group_fields, dropna=False, as_index=False).agg(
        gmv_usd=("gmv_usd", "sum"),
        orders=("orders", "sum"),
    )
    grouped["abs"] = grouped["gmv_usd"] / grouped["orders"].where(grouped["orders"] != 0)
    grouped["abs"] = grouped["abs"].fillna(0)
    return grouped


def _merge_yoy(current: pd.DataFrame, prior: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    prior = prior.rename(
        columns={
            "gmv_usd": "gmv_usd_prior_year",
            "orders": "orders_prior_year",
            "abs": "abs_prior_year",
        }
    )
    merged = current.merge(prior, on=keys, how="left")
    for field in ["gmv_usd_prior_year", "orders_prior_year", "abs_prior_year"]:
        merged[field] = merged[field].fillna(0)
    merged["gmv_yoy"] = _safe_growth(merged["gmv_usd"], merged["gmv_usd_prior_year"])
    merged["orders_yoy"] = _safe_growth(merged["orders"], merged["orders_prior_year"])
    merged["abs_change"] = merged["abs"] - merged["abs_prior_year"]
    return merged


def _safe_growth(current: pd.Series, prior: pd.Series) -> pd.Series:
    return ((current / prior.where(prior != 0)) - 1).fillna(0)


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
        sheet.freeze_panes = "A2"
