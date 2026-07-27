"""Assemble V2 Data Workbook from SolveLoop outputs."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from catemate.execution.result_collector import ExecutionResult
from catemate.orchestration.schemas import AnalysisPlan, ReportBlueprint, SolveLoopState, SolveVerdict
from catemate.schemas.data_workbook import (
    BlueprintSheetRow,
    DataWorkbookSpec,
    GapRow,
    PlanSheetRow,
    VerifyAuditRow,
)


def build_data_workbook_spec(
    *,
    blueprint: ReportBlueprint,
    plan: AnalysisPlan,
    verdict: SolveVerdict,
    execution: ExecutionResult | None = None,
) -> DataWorkbookSpec:
    blueprint_rows = [
        BlueprintSheetRow(
            section_id=s.section_id,
            title=s.title,
            sub_question=s.sub_question,
            presentation=s.expected_shape.presentation,
            scope_note=_blueprint_scope_note(plan, s.section_id),
        )
        for s in blueprint.sections
    ]
    plan_rows = [
        PlanSheetRow(
            run_id=r.run_id,
            section_id=r.section_id,
            module_id=r.module_id,
            metric_id=r.metric_id,
            grain=r.grain,
            is_sub_category=1 if r.is_sub_category else 0,
            scope_kind=r.scope_kind,
            source_kind=r.source_kind,
            table_id=r.table_id,
            status=r.status,
            scope_label=r.scope_label,
            missing=r.missing,
        )
        for r in plan.runs
    ]
    gap_rows = [
        GapRow(
            gap_id=f"gap_{u.section_id}",
            section_id=u.section_id,
            reason=u.reason,
            suggestion=u.suggestion,
        )
        for u in verdict.unsolved_sections
    ]
    verify_rows = [
        VerifyAuditRow(
            loop_iteration=verdict.loop_iteration,
            verdict=verdict.verdict,
            exit_reason=verdict.exit_reason or "",
            solved_sections=",".join(verdict.solved_sections),
            unsolved_sections=",".join(u.section_id for u in verdict.unsolved_sections),
        )
    ]
    _ = execution
    return DataWorkbookSpec(
        goal=blueprint.goal,
        blueprint_rows=blueprint_rows,
        plan_rows=plan_rows,
        gap_rows=gap_rows,
        verify_rows=verify_rows,
    )


def write_data_workbook(
    *,
    state: SolveLoopState,
    execution: ExecutionResult,
    output_path: Path,
) -> Path:
    if state.blueprint is None or state.plan is None or state.verdict is None:
        raise ValueError("SolveLoopState must include blueprint, plan, and verdict")

    spec = build_data_workbook_spec(
        blueprint=state.blueprint,
        plan=state.plan,
        verdict=state.verdict,
        execution=execution,
    )
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "Blueprint",
        ["section_id", "title", "sub_question", "presentation", "scope_note"],
        spec.blueprint_rows,
    )
    _write_sheet(
        wb,
        "Plan",
        [
            "run_id",
            "section_id",
            "module_id",
            "metric_id",
            "grain",
            "is_sub_category",
            "scope_kind",
            "source_kind",
            "table_id",
            "status",
            "scope_label",
            "missing",
        ],
        spec.plan_rows,
    )
    _write_sheet(wb, "Gaps", ["gap_id", "section_id", "reason", "suggestion"], spec.gap_rows)
    _write_sheet(
        wb,
        "Verify",
        ["loop_iteration", "verdict", "exit_reason", "solved_sections", "unsolved_sections"],
        spec.verify_rows,
    )

    sheet_names_by_key = workbook_sheet_names_by_storage_key(execution)
    for item in execution.tables:
        storage_key = item.get("storage_key")
        if not storage_key:
            continue
        df = execution.dataframes.get(storage_key)
        if df is None:
            continue
        sheet_name = sheet_names_by_key.get(str(storage_key))
        if not sheet_name:
            continue
        ws = wb.create_sheet(title=sheet_name)
        scope_note = df.attrs.get("scope_label")
        if scope_note:
            ws.append([f"scope: {scope_note}"])
        _write_dataframe(ws, df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([getattr(row, h) for h in headers])
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    note = _monthly_aggregation_note(df)
    if note:
        ws.append([note])
    header_row = ws.max_row + 1
    ws.append(list(df.columns))
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
    for record in df.itertuples(index=False, name=None):
        ws.append(list(record))


def data_sheet_name_for_table(*, section_id: str, table_id: str, run_id: str) -> str:
    """Public helper: expected Excel sheet title for an execution table row."""
    return _data_sheet_name(section_id=section_id, table_id=table_id, run_id=run_id)


def expected_data_sheet_names(execution: ExecutionResult) -> set[str]:
    """Return all Data.* sheet titles implied by execution.tables metadata."""
    return set(workbook_sheet_names_by_storage_key(execution).values())


def workbook_sheet_names_by_storage_key(execution: ExecutionResult) -> dict[str, str]:
    """Map execution storage_key to the Excel sheet title used in write_data_workbook."""
    used: set[str] = set()
    mapping: dict[str, str] = {}
    for item in execution.tables:
        storage_key = str(item.get("storage_key") or "")
        if not storage_key:
            continue
        section_id = str(item.get("section_id") or "section")
        table_id = str(item.get("table_id") or "table")
        run_id = str(item.get("run_id") or "")
        base = _data_sheet_name(section_id=section_id, table_id=table_id, run_id=run_id)
        mapping[storage_key] = _unique_sheet_name(base, table_id=table_id, used=used)
    return mapping


def _data_sheet_name(*, section_id: str, table_id: str, run_id: str) -> str:
    section_based = f"Data.{section_id}.{table_id}"
    if len(section_based) <= 31:
        return _safe_sheet_name(section_based)
    run_based = f"Data.{run_id}.{table_id}"
    return _safe_sheet_name(run_based)


def _unique_sheet_name(base: str, *, table_id: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base
    suffix = table_id.split("_")[-1][:4] or "t"
    candidate = _safe_sheet_name(f"{base[: max(1, 31 - 5)]}_{suffix}")
    index = 1
    while candidate in used:
        candidate = _safe_sheet_name(f"{base[: max(1, 28 - len(str(index)))]}_{index}")
        index += 1
    used.add(candidate)
    return candidate


def _safe_sheet_name(name: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in invalid else ch for ch in name)
    return cleaned[:31] or "Data"


def _monthly_aggregation_note(df: pd.DataFrame) -> str:
    columns = {str(col).strip().lower() for col in df.columns}
    if "grass_month" in columns or "month" in columns:
        return "注：本表为月度聚合数据（grass_month/month）。"
    return ""


def _blueprint_scope_note(plan: AnalysisPlan, section_id: str) -> str:
    for run in plan.runs:
        if run.section_id == section_id and run.scope_kind == "comparison":
            return "派生对比表（非 rawdata 源表）"
        if run.section_id == section_id and run.scope_kind == "subset":
            return "Plan 使用 item 粒度 + item_l3_category_csv（蓝图 grain 可为 category）"
    return ""
