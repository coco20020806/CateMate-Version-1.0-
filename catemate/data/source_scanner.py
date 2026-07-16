"""Scan raw data files and inspect expected workbook structure."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from catemate.schemas.category_requirement import FieldCheck, SheetCheck, SourceFile


def scan_excel_sources(raw_data_dir: Path) -> list[SourceFile]:
    """Return Excel files in raw data directory (flat or grain subdirs)."""
    files: list[SourceFile] = []
    search_paths: list[Path] = []
    for sub in ("category", "shop", "item"):
        grain_dir = raw_data_dir / sub
        if grain_dir.is_dir():
            search_paths.append(grain_dir)
    if not search_paths:
        search_paths = [raw_data_dir]

    for search_dir in search_paths:
        for path in sorted(search_dir.glob("*.xlsx")):
            stat = path.stat()
            grain = search_dir.name if search_dir.parent == raw_data_dir else None
            files.append(
                SourceFile(
                    path=path,
                    size_bytes=stat.st_size,
                    modified_time=datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                    matched_source_id=_guess_source_id(path.name, grain),
                )
            )
    return files


def _guess_source_id(filename: str, grain: str | None) -> str | None:
    name = filename.lower()
    if "sph" in name and "rm" in name:
        return "rm_raw_data"
    if "2026" in name or "品类数据看板" in filename:
        if grain == "shop":
            return "dashboard_top_shop"
        if grain == "item":
            return "dashboard_top_listing"
        return "dashboard_history"
    return None


def check_required_sheets(workbook_path: Path, required_sheets: list[str]) -> list[SheetCheck]:
    """Check whether expected sheets exist in a workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = set(workbook.sheetnames)
    return [
        SheetCheck(sheet_name=name, exists=name in sheet_names, note="" if name in sheet_names else "源文件中未找到该 sheet")
        for name in required_sheets
    ]


def check_raw_data_fields(workbook_path: Path, required_fields: list[str], sheet_name: str = "Raw data") -> list[FieldCheck]:
    """Check required fields in the first row of Raw data."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return [FieldCheck(field_name=field, exists=False, note=f"缺少 {sheet_name} sheet") for field in required_fields]

    sheet = workbook[sheet_name]
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {str(value).strip() for value in first_row if value is not None}
    return [
        FieldCheck(field_name=field, exists=field in headers, note="" if field in headers else "字段缺失")
        for field in required_fields
    ]
