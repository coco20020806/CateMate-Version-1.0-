"""Build searchable category tree lookup tables."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


TREE_HEADERS = [
    "source_row",
    "l1",
    "l2",
    "l3",
    "l4",
    "l1_zh",
    "l2_zh",
    "l3_zh",
    "l4_zh",
    "category_path",
    "category_path_zh",
]


def build_sph_category_tree_lookup(workbook_path: Path, output_path: Path) -> list[dict[str, str]]:
    """Convert the SPH category tree sheet into a flat lookup table."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook["SPH类目树"]
    rows: list[dict[str, str]] = []
    current = {key: "" for key in ["l1", "l2", "l3", "l4", "l1_zh", "l2_zh", "l3_zh", "l4_zh"]}

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx <= 2:
            continue
        values = list(row[:9]) + [None] * max(0, 9 - len(row[:9]))
        l1, l2, l3, l4 = values[0], values[1], values[2], values[3]
        z1, z2, z3, z4 = values[5], values[6], values[7], values[8]

        if l1:
            current.update({"l1": str(l1).strip(), "l2": "", "l3": "", "l4": ""})
        if l2:
            current.update({"l2": str(l2).strip(), "l3": "", "l4": ""})
        if l3:
            current.update({"l3": str(l3).strip(), "l4": ""})
        if l4:
            current["l4"] = str(l4).strip()

        if z1:
            current.update({"l1_zh": str(z1).strip(), "l2_zh": "", "l3_zh": "", "l4_zh": ""})
        if z2:
            current.update({"l2_zh": str(z2).strip(), "l3_zh": "", "l4_zh": ""})
        if z3:
            current.update({"l3_zh": str(z3).strip(), "l4_zh": ""})
        if z4:
            current["l4_zh"] = str(z4).strip()

        if any([l1, l2, l3, l4]):
            category_path = " > ".join(part for part in [current["l1"], current["l2"], current["l3"], current["l4"]] if part)
            category_path_zh = " > ".join(
                part for part in [current["l1_zh"], current["l2_zh"], current["l3_zh"], current["l4_zh"]] if part
            )
            rows.append(
                {
                    **current,
                    "source_row": str(row_idx),
                    "category_path": category_path,
                    "category_path_zh": category_path_zh,
                }
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=TREE_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    return rows


def find_category_candidates(lookup_rows: list[dict[str, str]], terms: list[str], limit: int = 5) -> list[dict[str, str]]:
    """Find simple exact/contains candidates in the category lookup."""
    candidates: list[dict[str, str]] = []
    seen: set[str] = set()
    for term in terms:
        term_norm = term.lower().strip()
        if not term_norm:
            continue
        for row in lookup_rows:
            haystack = f"{row['category_path']} {row['category_path_zh']}".lower()
            if term_norm in haystack and row["category_path"] not in seen:
                candidates.append({**row, "matched_text": term, "match_type": "contains"})
                seen.add(row["category_path"])
                if len([item for item in candidates if item["matched_text"] == term]) >= limit:
                    break
    return candidates
