"""Integration: workbook Data sheet names align with Plan sheet rows."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from catemate.core.paths import PROCESSED_DATA_DIR
from catemate.execution.runner import execute_analysis_plan
from catemate.modules.data_workbook import (
    expected_data_sheet_names,
    workbook_sheet_names_by_storage_key,
    write_data_workbook,
)
from catemate.orchestration.plan_composer import compose_analysis_plan
from catemate.orchestration.schemas import ReportBlueprint, SolveLoopState, SolveVerdict
from catemate.orchestration.module_registry import is_active_v2_module
from catemate.understanding.schemas import RequirementUnderstandingSpec

PROJECT_ROOT = Path(__file__).resolve().parents[3]
LATEST_RUN_DIR = PROJECT_ROOT / "outputs" / "runs" / "smart_pet_bowl_sg_20260716_154654"
UNDERSTANDING_PATH = LATEST_RUN_DIR / "requirement_understanding_smart_pet_bowl_sg_20260716_154654.json"
BLUEPRINT_PATH = LATEST_RUN_DIR / "report_blueprint_smart_pet_bowl_sg_20260716_154654.json"
VERDICT_PATH = LATEST_RUN_DIR / "solve_verdict_smart_pet_bowl_sg_20260716_154654.json"

PLAN_HEADERS = [
    "run_id",
    "section_id",
    "module_id",
    "metric_id",
    "grain",
    "is_sub_category",
    "scope_kind",
    "table_id",
    "status",
    "scope_label",
    "missing",
]

ACTIVE_MODULES = {"monthly_market_trend", "top_sku_info"}


def _load_latest_run_inputs() -> tuple[RequirementUnderstandingSpec, ReportBlueprint, SolveVerdict]:
    if not UNDERSTANDING_PATH.exists() or not BLUEPRINT_PATH.exists():
        pytest.skip("latest run artifacts unavailable")
    understanding = RequirementUnderstandingSpec.model_validate(
        json.loads(UNDERSTANDING_PATH.read_text(encoding="utf-8"))
    )
    blueprint = ReportBlueprint.model_validate(json.loads(BLUEPRINT_PATH.read_text(encoding="utf-8")))
    verdict = (
        SolveVerdict.model_validate(json.loads(VERDICT_PATH.read_text(encoding="utf-8")))
        if VERDICT_PATH.exists()
        else SolveVerdict(verdict="partial")
    )
    return understanding, blueprint, verdict


def _blueprint_with_active_modules_only(blueprint: ReportBlueprint) -> ReportBlueprint:
    sections = [
        section
        for section in blueprint.sections
        if section.module_id in ACTIVE_MODULES and is_active_v2_module(section.module_id)
    ]
    if not sections:
        raise ValueError("no active-module sections in latest run blueprint")
    return blueprint.model_copy(update={"sections": sections})


def _read_plan_rows(workbook_path: Path) -> list[dict[str, str]]:
    book = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Plan" not in book.sheetnames:
            raise AssertionError("workbook missing Plan sheet")
        ws = book["Plan"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        book.close()
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    assert headers == PLAN_HEADERS, f"unexpected Plan headers: {headers}"
    plan_rows: list[dict[str, str]] = []
    for raw in rows[1:]:
        if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        plan_rows.append(
            {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(raw)
            }
        )
    return plan_rows


def test_workbook_data_sheet_names_match_plan_for_latest_run_requirement(
    tmp_path: Path,
) -> None:
    """Latest run: 新加坡智能宠物碗 — Data sheet names must map to Plan rows."""
    if not PROCESSED_DATA_DIR.exists():
        pytest.skip("processed data unavailable")

    understanding, blueprint, verdict = _load_latest_run_inputs()
    request_text = understanding.original_request
    assert "新加坡" in request_text and "智能宠物碗" in request_text

    filtered_blueprint = _blueprint_with_active_modules_only(blueprint)
    plan = compose_analysis_plan(filtered_blueprint, understanding)
    state = SolveLoopState(
        blueprint=filtered_blueprint,
        plan=plan,
        verdict=verdict,
        phase="done",
    )

    execution = execute_analysis_plan(plan, processed_data_dir=PROCESSED_DATA_DIR)
    output_path = tmp_path / "data_workbook_test.xlsx"
    write_data_workbook(state=state, execution=execution, output_path=output_path)

    book = openpyxl.load_workbook(output_path, read_only=True)
    try:
        sheet_names = list(book.sheetnames)
    finally:
        book.close()

    meta_sheets = {"Blueprint", "Plan", "Gaps", "Verify"}
    data_sheets = [name for name in sheet_names if name not in meta_sheets]
    expected_sheets = expected_data_sheet_names(execution)

    assert expected_sheets, "execution produced no data tables"
    assert expected_sheets.issubset(set(sheet_names)), (
        f"missing data sheets: {sorted(expected_sheets - set(sheet_names))}"
    )
    assert set(data_sheets) == expected_sheets, (
        f"extra={sorted(set(data_sheets) - expected_sheets)}, "
        f"missing={sorted(expected_sheets - set(data_sheets))}"
    )

    plan_rows = _read_plan_rows(output_path)
    assert plan_rows, "Plan sheet has no rows"
    assert len(plan_rows) == len(plan.runs), "Plan sheet row count must match composed AnalysisPlan"

    plan_by_run_id = {row["run_id"]: row for row in plan_rows}
    for run in plan.runs:
        row = plan_by_run_id[run.run_id]
        assert row["section_id"] == run.section_id
        assert row["module_id"] == run.module_id
        assert row["metric_id"] == run.metric_id
        assert row["grain"] == run.grain
        assert row["is_sub_category"] == ("1" if run.is_sub_category else "0")
        assert row["scope_kind"] == run.scope_kind
        assert row["table_id"] == run.table_id

    sheet_by_key = workbook_sheet_names_by_storage_key(execution)

    for item in execution.tables:
        run_id = str(item.get("run_id") or "")
        plan_row = plan_by_run_id[run_id]
        storage_key = str(item.get("storage_key") or "")
        expected_name = sheet_by_key[storage_key]
        assert expected_name in sheet_names
        assert plan_row["section_id"] == str(item.get("section_id") or "")
        assert plan_row["metric_id"] == str(item.get("metric_id") or "")

    for row in plan_rows:
        if row["status"] != "executable":
            continue
        run_tables = [item for item in execution.tables if str(item.get("run_id") or "") == row["run_id"]]
        assert run_tables, f"executable Plan row {row['run_id']} has no execution tables"
        for item in run_tables:
            storage_key = str(item.get("storage_key") or "")
            assert sheet_by_key[storage_key] in sheet_names

    subset_rows = [row for row in plan_rows if row.get("is_sub_category") == "1"]
    assert subset_rows, "latest run is sub-L3; expect subset rows in Plan"
    for row in subset_rows:
        assert row["grain"] == "item"
        if row["module_id"] == "monthly_market_trend":
            assert row["table_id"] == "item_l3_category_csv"
