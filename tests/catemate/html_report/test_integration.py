"""Integration test for html_report propose + render flow."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import Workbook

from catemate.html_report.generator import propose_visual_report, render_html_report_from_spec
from catemate.html_report.proposal_generator import load_visual_report_spec, save_visual_report_spec


def _write_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data.s1.orders_by_site_month")
    ws.append(["grass_region", "grass_month", "orders"])
    ws.append(["BR", "2026-05-01", 54000.0])
    ws.append(["BR", "2026-06-01", 58000.0])
    ws2 = wb.create_sheet("Data.s1.orders_latest_month_by_site")
    ws2.append(["grass_region", "orders"])
    ws2.append(["BR", 54000.0])
    wb.save(path)


def _mock_llm_payload() -> dict:
    return {
        "case_id": "demo",
        "original_question": "orders trend?",
        "report_goal": "Analyze orders",
        "executive_summary": "BR orders 54k.",
        "sections": [
            {
                "section_id": "s1",
                "title": "Orders trend",
                "sub_question": "Trend?",
                "narrative": "BR latest 54k.",
                "status": "solved",
                "charts": [
                    {
                        "chart_id": "s1_orders_by_site_month",
                        "section_id": "s1",
                        "table_id": "orders_by_site_month",
                        "module_id": "monthly_market_trend",
                        "chart_type": "trend",
                        "title": "Orders by site month",
                        "x_field": "grass_month",
                        "y_fields": ["orders"],
                        "series_field": "grass_region",
                        "visible": True,
                        "role": "primary",
                        "binding_source": "llm",
                        "confidence": "high",
                        "notes": [],
                    },
                    {
                        "chart_id": "s1_orders_latest_month_by_site",
                        "section_id": "s1",
                        "table_id": "orders_latest_month_by_site",
                        "module_id": "monthly_market_trend",
                        "chart_type": "bar",
                        "title": "Latest month",
                        "x_field": "grass_region",
                        "y_fields": ["orders"],
                        "visible": True,
                        "role": "secondary",
                        "binding_source": "llm",
                        "confidence": "high",
                        "notes": [],
                    },
                ],
            }
        ],
        "data_gaps": [],
        "generated_at": "2026-07-17T00:00:00+00:00",
        "spec_status": "draft",
    }


def test_propose_confirm_render_integration(tmp_path: Path) -> None:
    wb = tmp_path / "data_workbook_demo.xlsx"
    spec_path = tmp_path / "visual_report_spec_demo.json"
    html_path = tmp_path / "html_report_demo.html"
    _write_workbook(wb)

    mock_client = MagicMock()
    mock_client.complete_json.return_value = _mock_llm_payload()

    propose_visual_report(
        workbook_path=wb,
        original_question="orders trend?",
        case_id="demo",
        spec_output=spec_path,
        ai_client=mock_client,
    )
    spec = load_visual_report_spec(spec_path)
    assert spec.spec_status == "draft"

    confirmed = spec.model_copy(update={"spec_status": "confirmed"})
    save_visual_report_spec(confirmed, spec_path)

    out = render_html_report_from_spec(
        spec=spec_path,
        workbook_path=wb,
        html_output=html_path,
    )
    assert out.exists()
    html = out.read_text(encoding="utf-8")
    assert "Analyze orders" in html
    assert "plotly" in html.lower()
