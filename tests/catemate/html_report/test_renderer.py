"""Tests for HTML report renderer."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from catemate.html_report.renderer import render_html_report
from catemate.html_report.schemas import ChartBinding, VisualReportSection, VisualReportSpec


def _write_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data.s_orders.orders_by_site_month")
    ws.append(["grass_region", "grass_month", "orders"])
    ws.append(["BR", "2026-05-01", 54000.0])
    ws.append(["BR", "2026-06-01", 58000.0])
    ws2 = wb.create_sheet("Data.s_orders.orders_latest_month_by_site")
    ws2.append(["grass_region", "orders"])
    ws2.append(["BR", 54000.0])
    wb.save(path)


def _confirmed_spec() -> VisualReportSpec:
    return VisualReportSpec(
        case_id="demo",
        original_question="orders trend?",
        report_goal="Analyze orders",
        executive_summary="BR orders trending.",
        spec_status="confirmed",
        sections=[
            VisualReportSection(
                section_id="s_orders",
                title="Orders trend",
                narrative="BR latest 54k.",
                status="solved",
                charts=[
                    ChartBinding(
                        chart_id="c_trend",
                        section_id="s_orders",
                        table_id="orders_by_site_month",
                        chart_type="trend",
                        title="Orders trend",
                        x_field="grass_month",
                        y_fields=["orders"],
                        series_field="grass_region",
                    ),
                    ChartBinding(
                        chart_id="c_bar",
                        section_id="s_orders",
                        table_id="orders_latest_month_by_site",
                        chart_type="bar",
                        title="Latest month",
                        x_field="grass_region",
                        y_fields=["orders"],
                        role="secondary",
                    ),
                ],
            )
        ],
    )


def test_render_html_report_writes_file(tmp_path: Path) -> None:
    wb = tmp_path / "data_workbook_demo.xlsx"
    out = tmp_path / "html_report_demo.html"
    _write_workbook(wb)

    result = render_html_report(
        spec=_confirmed_spec(),
        workbook_path=wb,
        output_path=out,
    )
    assert result.exists()
    html = result.read_text(encoding="utf-8")
    assert "Analyze orders" in html
    assert "plotly" in html.lower()
    assert "Orders trend" in html
    assert "显示数字" in html
    assert "2026-05" in html
    assert "2026-06" in html
