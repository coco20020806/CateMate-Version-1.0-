"""Tests for workbook data loader metadata handling."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from catemate.html_report.data_loader import dataframe_from_sheet_rows, load_workbook_table_entries


def _write_metadata_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data.r2.gmv_by_site_month")
    ws.append(["scope: ALL / Pets"])
    ws.append(["注：本表为月度聚合数据（grass_month/month）。"])
    ws.append(["grass_region", "grass_month", "gmv_usd"])
    ws.append(["BR", "2026-05-01", 100.0])
    wb.save(path)


def test_dataframe_from_sheet_rows_skips_metadata() -> None:
    rows = [
        ("scope: ALL / Pets", None, None),
        ("注：本表为月度聚合数据（grass_month/month）。", None, None),
        ("grass_region", "grass_month", "gmv_usd"),
        ("BR", "2026-05-01", 100.0),
    ]
    df = dataframe_from_sheet_rows(rows)
    assert list(df.columns) == ["grass_region", "grass_month", "gmv_usd"]
    assert len(df) == 1
    assert float(df.iloc[0]["gmv_usd"]) == 100.0


def test_load_workbook_table_entries_parses_realistic_sheet(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _write_metadata_workbook(wb)
    entries = load_workbook_table_entries(wb)
    assert len(entries) == 1
    assert list(entries[0].df.columns) == ["grass_region", "grass_month", "gmv_usd"]
