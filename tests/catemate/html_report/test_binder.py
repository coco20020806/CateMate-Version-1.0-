"""Tests for rule-based chart binder."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from catemate.html_report.binder import build_draft_bindings
from catemate.orchestration.schemas import AnalysisPlan, BlueprintSection, ExpectedShape, PlanRun, ReportBlueprint


def _write_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data.s1.orders_by_site_month")
    ws.append(["grass_region", "grass_month", "orders"])
    ws.append(["BR", "2026-05-01", 54000.0])
    ws.append(["BR", "2026-06-01", 58000.0])
    ws2 = wb.create_sheet("Data.s1.orders_latest_site")
    ws2.append(["grass_region", "orders"])
    ws2.append(["BR", 54000.0])
    wb.save(path)


def _write_plan(path: Path) -> None:
    plan = AnalysisPlan(
        goal="orders trend",
        runs=[
            PlanRun(
                run_id="r1",
                section_id="s_orders",
                module_id="monthly_market_trend",
                metric_id="orders",
                table_id="orders_by_site_month",
                status="executed",
            ),
            PlanRun(
                run_id="r2",
                section_id="s_orders",
                module_id="monthly_market_trend",
                metric_id="orders",
                table_id="orders_latest_site",
                status="executed",
            ),
        ],
    )
    path.write_text(json.dumps(plan.model_dump(mode="json"), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_blueprint(path: Path) -> None:
    blueprint = ReportBlueprint(
        goal="Analyze orders",
        sections=[
            BlueprintSection(
                section_id="s_orders",
                title="Orders trend",
                sub_question="How do orders trend?",
                expected_shape=ExpectedShape(presentation="trend", metrics=["orders"]),
                module_id="monthly_market_trend",
                metric_id="orders",
            )
        ],
    )
    path.write_text(json.dumps(blueprint.model_dump(mode="json"), ensure_ascii=False, indent=2), encoding="utf-8")


def test_binder_assigns_primary_and_secondary_roles(tmp_path: Path) -> None:
    wb = tmp_path / "data_workbook_demo.xlsx"
    plan = tmp_path / "plan.json"
    blueprint = tmp_path / "blueprint.json"
    _write_workbook(wb)
    _write_plan(plan)
    _write_blueprint(blueprint)

    draft = build_draft_bindings(
        workbook_path=wb,
        original_question="orders trend?",
        blueprint_path=blueprint,
        plan_path=plan,
    )
    assert len(draft.bindings) == 2
    by_table = {b.table_id: b for b in draft.bindings}
    assert by_table["orders_by_site_month"].chart_type == "trend"
    assert by_table["orders_by_site_month"].role == "primary"
    assert by_table["orders_latest_site"].chart_type == "bar"
    assert by_table["orders_latest_site"].role == "secondary"
    assert draft.sections[0].section_id == "s_orders"


def test_binder_uses_chart_preset_when_available(tmp_path: Path) -> None:
    wb = tmp_path / "data_workbook_demo.xlsx"
    _write_workbook(wb)
    draft = build_draft_bindings(
        workbook_path=wb,
        original_question="orders trend?",
    )
    trend = next(b for b in draft.bindings if b.table_id == "orders_by_site_month")
    assert trend.binding_source in {"chart_preset", "heuristic", "blueprint"}
    assert trend.y_fields
