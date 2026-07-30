"""Composer / renderer tests for print_vertical_report."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from catemate.html_report.schemas import ChartBinding, VisualReportSection, VisualReportSpec
from catemate.print_report.composer import compose_print_report_doc
from catemate.print_report.generator import generate_print_report
from catemate.print_report.renderer import render_print_report_html


def _write_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data.s_market.gmv_by_site_month")
    ws.append(["grass_region", "grass_month", "gmv_usd"])
    ws.append(["BR", "2026-05-01", 54000.0])
    ws.append(["BR", "2026-06-01", 58000.0])
    ws.append(["PH", "2026-05-01", 12000.0])
    ws.append(["PH", "2026-06-01", 15000.0])
    ws2 = wb.create_sheet("Data.s_market.gmv_latest_month_by_site")
    ws2.append(["grass_region", "gmv_usd"])
    ws2.append(["BR", 58000.0])
    ws2.append(["PH", 15000.0])
    wb.save(path)


def _confirmed_spec() -> VisualReportSpec:
    return VisualReportSpec(
        case_id="demo_print",
        original_question="智能喂食器趋势？",
        report_goal="智能喂食器类目深度分析",
        executive_summary="BR 体量领先，PH 值得跟进验证。",
        spec_status="confirmed",
        sections=[
            VisualReportSection(
                section_id="s_market",
                title="市场趋势",
                narrative="BR 为主力市场。",
                status="solved",
                charts=[
                    ChartBinding(
                        chart_id="c_trend",
                        section_id="s_market",
                        table_id="gmv_by_site_month",
                        chart_type="trend",
                        title="GMV 月度趋势",
                        x_field="grass_month",
                        y_fields=["gmv_usd"],
                        series_field="grass_region",
                    ),
                    ChartBinding(
                        chart_id="c_bar",
                        section_id="s_market",
                        table_id="gmv_latest_month_by_site",
                        chart_type="bar",
                        title="最新月各站",
                        x_field="grass_region",
                        y_fields=["gmv_usd"],
                        role="secondary",
                    ),
                ],
            )
        ],
    )


def test_compose_requires_confirmed_spec(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _write_workbook(wb)
    draft = _confirmed_spec().model_copy(update={"spec_status": "draft"})
    with pytest.raises(ValueError, match="confirmed"):
        compose_print_report_doc(spec=draft, workbook_path=wb)


def test_compose_print_report_page_order_and_fuzzy(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _write_workbook(wb)
    doc = compose_print_report_doc(spec=_confirmed_spec(), workbook_path=wb)
    assert doc.fuzzy_applied is True
    layouts = [page.layout for page in doc.pages]
    assert layouts[0] == "cover"
    assert layouts[1] == "toc"
    assert layouts[2] == "insight"
    assert "evidence" in layouts
    assert layouts[-1] == "actions"
    assert doc.next_actions
    # No raw large exact money values in fuzzy metric displays.
    for page in doc.pages:
        for card in page.cards:
            for metric in card.fuzzy_metrics:
                assert "54000" not in metric.display
                assert "58000" not in metric.display
                assert metric.raw_suppressed is True or metric.kind == "label"


def test_render_print_html_contains_brand_and_editable(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _write_workbook(wb)
    outputs = generate_print_report(
        spec=_confirmed_spec(),
        workbook_path=wb,
        html_output=tmp_path / "print.html",
        doc_output=tmp_path / "print.json",
        case_id="demo_print",
        timestamp="20260729",
    )
    html = outputs.html_path.read_text(encoding="utf-8")
    assert "report-page" in html
    assert 'contenteditable="true"' in html
    assert "--brand:" in html
    assert "#EE4D2D" in html
    assert "模糊化" in html
    assert "54000" not in html
    assert "58,000" not in html
    assert outputs.doc_path.exists()
